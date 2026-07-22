"""
Excel 报表生成模块 (rebalance_report.py)
======================================
职责：
  write_rebalance_day_report — 将调仓日报表写入单文件 Excel（含全部 Sheet）

导出：
  write_rebalance_day_report
"""

from __future__ import annotations

import os
import sys
from datetime import datetime
from typing import Optional

import numpy as np
import pandas as pd
import pandas_market_calendars as pmc

from qqq_core.performance_metrics import (
    performance_summary,
    worst_period_drawdown,
)
from data.price_snapshot import manifest_adjustments_frame

# ── 路径注册（strategy_utils 位于同级目录）────────────────────────────────────
_HERE = os.path.dirname(os.path.abspath(__file__))
_PARENT = os.path.dirname(_HERE)
if _PARENT not in sys.path:
    sys.path.insert(0, _PARENT)

from strategy_utils import filter_weight_lt
from tp_sl_exit import (
    EXIT_DYNAMIC_TP_SL,
    EXIT_STOP_LOSS,
    EXIT_TAKE_PROFIT,
    thresholds_for_day,
)


WEIGHT_FILTER_THRESHOLD: float = 0.0001
ALL_OPERATIONS_WEIGHT_FILTER_THRESHOLD: float = 0.01
TP_SL_ACTION_LOOKAHEAD_DATES: int = 5
PERIOD_SUMMARY_2_START_DATE = "2026-03-27"
RETURN_ATTRIBUTION_RECENT_YEARS: int = 2


def _describe_composite_method(sheet_name: str) -> str:
    """给报告用的复合方法说明；m1 明确标注全样本前瞻偏误。"""
    if "_m1" in sheet_name:
        return f"{sheet_name}（全样本 oracle baseline，含前瞻偏误，仅供研究对比）"
    if sheet_name.startswith("ic_"):
        return f"{sheet_name}（IC 加权）"
    if sheet_name.startswith("rank_ic_"):
        return f"{sheet_name}（Rank IC 加权）"
    if sheet_name.startswith("beta_"):
        return f"{sheet_name}（Beta 加权）"
    if sheet_name.startswith("ols_"):
        return f"{sheet_name}（OLS 多元回归加权）"
    if sheet_name.startswith("pca_"):
        return f"{sheet_name}（PCA 主成分）"
    if sheet_name.startswith("rank_"):
        return f"{sheet_name}（截面排名复合）"
    return sheet_name


def build_period_summary_2(
    period_summary: pd.DataFrame,
    start_date: str | pd.Timestamp = PERIOD_SUMMARY_2_START_DATE,
) -> pd.DataFrame:
    """Return the period summary from ``start_date`` with cumulative return reset."""

    if period_summary.empty:
        return period_summary.copy()

    required_columns = {"Rebalance_Date", "Period_Return"}
    missing = required_columns.difference(period_summary.columns)
    if missing:
        raise KeyError(f"Period summary missing required columns: {sorted(missing)}")

    rebalance_dates = pd.to_datetime(period_summary["Rebalance_Date"], errors="coerce")
    summary = period_summary.loc[
        rebalance_dates >= pd.Timestamp(start_date).normalize()
    ].copy()
    summary.reset_index(drop=True, inplace=True)

    if summary.empty:
        return summary

    period_returns = pd.to_numeric(summary["Period_Return"], errors="coerce")
    summary["Period_Cumulative_Return"] = (1.0 + period_returns).cumprod() - 1.0
    return summary


def _as_naive_trading_day(value) -> pd.Timestamp:
    """Normalize NYSE calendar values to timezone-naive midnight timestamps."""

    ts = pd.Timestamp(value)
    if ts.tzinfo is not None:
        ts = ts.tz_convert(None)
    return ts.normalize()


def _clean_dated_returns(returns: pd.Series) -> pd.Series:
    """Return a finite numeric return series with a normalized datetime index."""

    if returns is None or len(returns) == 0:
        return pd.Series(dtype=float)

    cleaned = pd.Series(returns).copy()
    cleaned.index = pd.to_datetime(cleaned.index, errors="coerce")
    cleaned = pd.to_numeric(cleaned, errors="coerce")
    valid = (~cleaned.index.isna()) & np.isfinite(cleaned.to_numpy(dtype=float))
    cleaned = cleaned.loc[valid].astype(float).sort_index()
    if cleaned.index.tz is not None:
        cleaned.index = cleaned.index.tz_convert(None)
    cleaned.index = cleaned.index.normalize()
    return cleaned


def build_performance_by_year(
    daily_returns: pd.Series,
    rf_rate: float = 0.02,
) -> pd.DataFrame:
    """Build calendar-year performance plus a flat-year CAGR stress test."""

    columns = [
        "Year",
        "Is_Partial_Year",
        "Total_Return",
        "Annual_Volatility",
        "Sharpe",
        "Max_Drawdown",
        "Full_CAGR_If_Year_Flat",
    ]
    returns = _clean_dated_returns(daily_returns)
    if returns.empty:
        return pd.DataFrame(columns=columns)

    nyse = pmc.get_calendar("NYSE")
    records: list[dict] = []
    for year in sorted(returns.index.year.unique()):
        year_mask = returns.index.year == int(year)
        year_returns = returns.loc[year_mask]
        summary = performance_summary(year_returns, rf=rf_rate, periods_per_year=252)

        expected_days = {
            _as_naive_trading_day(day)
            for day in nyse.valid_days(f"{int(year)}-01-01", f"{int(year)}-12-31")
        }
        observed_days = set(pd.DatetimeIndex(year_returns.index).normalize())

        flat_returns = returns.copy()
        flat_returns.loc[year_mask] = 0.0
        flat_summary = performance_summary(flat_returns, rf=rf_rate, periods_per_year=252)

        records.append(
            {
                "Year": int(year),
                "Is_Partial_Year": observed_days != expected_days,
                "Total_Return": summary["total_return"],
                "Annual_Volatility": summary["annual_vol"],
                "Sharpe": summary["sharpe"],
                "Max_Drawdown": summary["max_drawdown"],
                "Full_CAGR_If_Year_Flat": flat_summary["annual_return"],
            }
        )

    return pd.DataFrame(records, columns=columns)


def build_period_return_attribution(period_summary: pd.DataFrame) -> pd.DataFrame:
    """Rank rebalance periods and express their additive log-return contribution."""

    columns = [
        "Rebalance_Date",
        "Next_Rebalance_Date",
        "Holding_Days",
        "Period_Return",
        "Log_Return_Contribution",
        "Net_Log_Contribution_Share",
        "Positive_Log_Contribution_Share",
        "Return_Rank",
        "Symbols",
    ]
    required = {"Rebalance_Date", "Period_Return"}
    if period_summary is None or period_summary.empty or not required.issubset(period_summary.columns):
        return pd.DataFrame(columns=columns)

    frame = period_summary.copy()
    frame["Rebalance_Date"] = pd.to_datetime(frame["Rebalance_Date"], errors="coerce")
    if "Next_Rebalance_Date" not in frame.columns:
        frame["Next_Rebalance_Date"] = pd.NaT
    else:
        frame["Next_Rebalance_Date"] = pd.to_datetime(
            frame["Next_Rebalance_Date"], errors="coerce"
        )
    if "Holding_Days" not in frame.columns:
        frame["Holding_Days"] = np.nan
    if "Symbols" not in frame.columns:
        frame["Symbols"] = ""

    frame["Period_Return"] = pd.to_numeric(frame["Period_Return"], errors="coerce")
    valid_log = frame["Period_Return"] > -1.0
    frame["Log_Return_Contribution"] = np.nan
    frame.loc[valid_log, "Log_Return_Contribution"] = np.log1p(
        frame.loc[valid_log, "Period_Return"]
    )

    total_log = frame["Log_Return_Contribution"].sum(min_count=1)
    positive_log = frame["Log_Return_Contribution"].clip(lower=0.0)
    positive_total = positive_log.sum(min_count=1)
    frame["Net_Log_Contribution_Share"] = (
        frame["Log_Return_Contribution"] / total_log
        if pd.notna(total_log) and abs(float(total_log)) > 1e-12
        else np.nan
    )
    frame["Positive_Log_Contribution_Share"] = (
        positive_log / positive_total
        if pd.notna(positive_total) and float(positive_total) > 1e-12
        else np.nan
    )
    frame["Return_Rank"] = (
        frame["Period_Return"].rank(method="min", ascending=False, na_option="bottom").astype("Int64")
    )
    return frame[columns].sort_values("Return_Rank", na_position="last").reset_index(drop=True)


def build_ticker_return_attribution(
    operations: pd.DataFrame,
    sample_end: pd.Timestamp,
    recent_years: int = RETURN_ATTRIBUTION_RECENT_YEARS,
) -> tuple[pd.DataFrame, pd.Timestamp]:
    """Aggregate holding frequency, weights, and simple return contribution by ticker."""

    columns = [
        "Symbol",
        "Holding_Count",
        "Recent_Holding_Count",
        "Average_Weight",
        "Average_Period_Return",
        "Best_Period_Return",
        "Worst_Period_Return",
        "Simple_Return_Contribution",
        "Recent_Simple_Return_Contribution",
        "Recent_Contribution_Share",
        "Contribution_Rank",
        "Recent_Contribution_Rank",
    ]
    cutoff = pd.Timestamp(sample_end).normalize() - pd.DateOffset(years=int(recent_years))
    required = {"Symbol", "Weight", "Period_Return", "Rebalance_Date"}
    if operations is None or operations.empty or not required.issubset(operations.columns):
        return pd.DataFrame(columns=columns), cutoff

    ops = operations.copy()
    ops["Symbol"] = ops["Symbol"].astype(str)
    ops["Weight"] = pd.to_numeric(ops["Weight"], errors="coerce")
    ops["Period_Return"] = pd.to_numeric(ops["Period_Return"], errors="coerce")
    ops["Rebalance_Date"] = pd.to_datetime(ops["Rebalance_Date"], errors="coerce")
    if "Next_Rebalance_Date" in ops.columns:
        ops["Next_Rebalance_Date"] = pd.to_datetime(
            ops["Next_Rebalance_Date"], errors="coerce"
        )
        recent_mask = ops["Next_Rebalance_Date"] >= cutoff
    else:
        recent_mask = ops["Rebalance_Date"] >= cutoff

    ops = ops.dropna(subset=["Symbol", "Weight", "Period_Return", "Rebalance_Date"])
    ops = ops[ops["Weight"] >= WEIGHT_FILTER_THRESHOLD].copy()
    if ops.empty:
        return pd.DataFrame(columns=columns), cutoff
    ops["Simple_Return_Contribution"] = ops["Weight"] * ops["Period_Return"]
    ops["Recent_Simple_Return_Contribution"] = np.where(
        recent_mask.reindex(ops.index, fill_value=False),
        ops["Simple_Return_Contribution"],
        0.0,
    )
    ops["Recent_Holding"] = recent_mask.reindex(ops.index, fill_value=False).astype(int)

    grouped = ops.groupby("Symbol", sort=True)
    result = grouped.agg(
        Holding_Count=("Symbol", "size"),
        Recent_Holding_Count=("Recent_Holding", "sum"),
        Average_Weight=("Weight", "mean"),
        Average_Period_Return=("Period_Return", "mean"),
        Best_Period_Return=("Period_Return", "max"),
        Worst_Period_Return=("Period_Return", "min"),
        Simple_Return_Contribution=("Simple_Return_Contribution", "sum"),
        Recent_Simple_Return_Contribution=("Recent_Simple_Return_Contribution", "sum"),
    ).reset_index()

    recent_total = result["Recent_Simple_Return_Contribution"].sum()
    result["Recent_Contribution_Share"] = (
        result["Recent_Simple_Return_Contribution"] / recent_total
        if abs(float(recent_total)) > 1e-12
        else np.nan
    )
    result["Contribution_Rank"] = (
        result["Simple_Return_Contribution"].rank(method="min", ascending=False).astype("Int64")
    )
    result["Recent_Contribution_Rank"] = (
        result["Recent_Simple_Return_Contribution"]
        .rank(method="min", ascending=False)
        .astype("Int64")
    )
    result = result[columns].sort_values("Contribution_Rank").reset_index(drop=True)
    return result, cutoff


def _daily_ticker_contributions(
    daily_returns: pd.Series,
    period_summary: pd.DataFrame,
    operations: pd.DataFrame,
    stock_returns: pd.DataFrame,
    exit_policy: str,
) -> pd.DataFrame:
    """Reconstruct each held ticker's contribution to the reported daily return series."""

    returns = _clean_dated_returns(daily_returns)
    if returns.empty or period_summary is None or operations is None or stock_returns is None:
        return pd.DataFrame(index=returns.index)
    required_period = {"Rebalance_Date", "Next_Rebalance_Date"}
    required_ops = {"Rebalance_Date", "Symbol", "Weight"}
    if (
        period_summary.empty
        or operations.empty
        or stock_returns.empty
        or not required_period.issubset(period_summary.columns)
        or not required_ops.issubset(operations.columns)
    ):
        return pd.DataFrame(index=returns.index)

    periods = period_summary.copy()
    periods["Rebalance_Date"] = pd.to_datetime(periods["Rebalance_Date"], errors="coerce")
    periods["Next_Rebalance_Date"] = pd.to_datetime(
        periods["Next_Rebalance_Date"], errors="coerce"
    )
    ops = operations.copy()
    ops["Rebalance_Date"] = pd.to_datetime(ops["Rebalance_Date"], errors="coerce")
    ops["Symbol"] = ops["Symbol"].astype(str)
    ops["Weight"] = pd.to_numeric(ops["Weight"], errors="coerce")
    if "Exit_Date" in ops.columns:
        ops["Exit_Date"] = pd.to_datetime(ops["Exit_Date"], errors="coerce")

    stock = stock_returns.copy()
    stock.index = pd.to_datetime(stock.index, errors="coerce")
    stock = stock.loc[~stock.index.isna()].sort_index()
    if stock.index.tz is not None:
        stock.index = stock.index.tz_convert(None)
    stock.index = stock.index.normalize()
    stock.columns = stock.columns.astype(str)

    held_symbols = sorted(set(ops["Symbol"]).intersection(stock.columns))
    contributions = pd.DataFrame(0.0, index=returns.index, columns=held_symbols)
    sample_end = returns.index.max()

    for _, period in periods.iterrows():
        rb_date = period["Rebalance_Date"]
        next_date = period["Next_Rebalance_Date"]
        if pd.isna(rb_date):
            continue
        rb_date = pd.Timestamp(rb_date).normalize()
        period_end = sample_end if pd.isna(next_date) else min(pd.Timestamp(next_date).normalize(), sample_end)
        period_dates = returns.index[(returns.index > rb_date) & (returns.index <= period_end)]
        if len(period_dates) == 0:
            continue

        period_ops = ops[ops["Rebalance_Date"].dt.normalize() == rb_date].copy()
        period_ops = period_ops.dropna(subset=["Symbol", "Weight"])
        period_ops = period_ops[period_ops["Weight"] >= WEIGHT_FILTER_THRESHOLD]
        if period_ops.empty:
            continue
        weights = period_ops.groupby("Symbol")["Weight"].sum()
        symbols = [symbol for symbol in weights.index if symbol in stock.columns]
        if not symbols:
            continue
        weights = weights.reindex(symbols)
        period_stock = stock.reindex(period_dates)[symbols]

        if str(exit_policy) == EXIT_DYNAMIC_TP_SL:
            period_contrib = period_stock.fillna(0.0).mul(weights, axis=1)
            if "Exit_Date" in period_ops.columns:
                exit_dates = period_ops.groupby("Symbol")["Exit_Date"].max()
                for symbol in symbols:
                    exit_date = exit_dates.get(symbol, pd.NaT)
                    if pd.notna(exit_date):
                        period_contrib.loc[
                            period_contrib.index > pd.Timestamp(exit_date).normalize(), symbol
                        ] = 0.0
        else:
            valid_mask = period_stock.notna()
            daily_weight_sum = valid_mask.mul(weights, axis=1).sum(axis=1)
            valid_days = daily_weight_sum > 1e-12
            normalized_weights = (
                valid_mask.mul(weights, axis=1)
                .div(daily_weight_sum.where(valid_days), axis=0)
                .fillna(0.0)
            )
            period_contrib = period_stock.fillna(0.0) * normalized_weights

        contributions.loc[period_dates, symbols] = (
            contributions.loc[period_dates, symbols].to_numpy()
            + period_contrib.reindex(period_dates).fillna(0.0).to_numpy()
        )

    return contributions


def build_ticker_exclusion_stress(
    daily_returns: pd.Series,
    period_summary: pd.DataFrame,
    operations: pd.DataFrame,
    stock_returns: pd.DataFrame,
    exit_policy: str,
    rf_rate: float = 0.02,
) -> pd.DataFrame:
    """Calculate CAGR after replacing each held ticker's daily contribution with cash."""

    columns = [
        "Symbol",
        "Full_CAGR",
        "CAGR_If_Excluded_To_Cash",
        "CAGR_Difference",
        "Impact_Rank",
        "Data_Status",
    ]
    returns = _clean_dated_returns(daily_returns)
    contributions = _daily_ticker_contributions(
        returns,
        period_summary,
        operations,
        stock_returns,
        exit_policy,
    )
    if returns.empty or contributions.empty:
        return pd.DataFrame(columns=columns)

    full_cagr = performance_summary(returns, rf=rf_rate, periods_per_year=252)[
        "annual_return"
    ]
    records = []
    for symbol in contributions.columns:
        excluded_returns = returns - contributions[symbol].reindex(returns.index).fillna(0.0)
        excluded_cagr = performance_summary(
            excluded_returns, rf=rf_rate, periods_per_year=252
        )["annual_return"]
        records.append(
            {
                "Symbol": symbol,
                "Full_CAGR": full_cagr,
                "CAGR_If_Excluded_To_Cash": excluded_cagr,
                "CAGR_Difference": full_cagr - excluded_cagr,
                "Data_Status": "OK",
            }
        )

    result = pd.DataFrame(records)
    result["Impact_Rank"] = (
        result["CAGR_Difference"].rank(method="min", ascending=False).astype("Int64")
    )
    return result[columns].sort_values("Impact_Rank").reset_index(drop=True)


def build_return_attribution(
    daily_returns: pd.Series,
    period_summary: pd.DataFrame,
    operations: pd.DataFrame,
    stock_returns: pd.DataFrame,
    exit_policy: str,
    rf_rate: float = 0.02,
    recent_years: int = RETURN_ATTRIBUTION_RECENT_YEARS,
) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    """Build all Return_Attribution sheet sections and methodology notes."""

    returns = _clean_dated_returns(daily_returns)
    sample_end = returns.index.max() if not returns.empty else pd.Timestamp(datetime.now().date())
    period_attr = build_period_return_attribution(period_summary)
    ticker_attr, cutoff = build_ticker_return_attribution(
        operations, sample_end, recent_years=recent_years
    )
    exclusion = build_ticker_exclusion_stress(
        returns,
        period_summary,
        operations,
        stock_returns,
        exit_policy,
        rf_rate=rf_rate,
    )
    methodology = pd.DataFrame(
        [
            ["Recent_Window_Years", int(recent_years)],
            ["Recent_Cutoff_Date", pd.Timestamp(cutoff).normalize()],
            ["Recent_Period_Rule", "Next_Rebalance_Date >= cutoff"],
            ["Holding_Filter", f"Weight >= {WEIGHT_FILTER_THRESHOLD:.4f}"],
            ["Period_Contribution", "log1p(Period_Return); additive through time"],
            ["Ticker_Simple_Contribution", "Weight * holding-period stock return"],
            [
                "Ticker_Exclusion_Method",
                "Replace the ticker's reconstructed daily contribution with cash; keep other weights unchanged",
            ],
            ["Transaction_Cost_Treatment", "Keep the original portfolio transaction cost unchanged"],
            ["CAGR_Annualization", "252 trading days"],
            [
                "Reconciliation_Note",
                "Simple ticker contributions are diagnostic and need not equal compounded portfolio returns",
            ],
        ],
        columns=["Methodology_Item", "Definition"],
    )
    return period_attr, ticker_attr, exclusion, methodology


def _style_table_header(worksheet, row_number: int, max_column: int) -> None:
    """Apply a compact report-table header style to one worksheet row."""

    from openpyxl.styles import Alignment, Font, PatternFill

    fill = PatternFill("solid", fgColor="1F4E78")
    font = Font(color="FFFFFF", bold=True)
    for column in range(1, max_column + 1):
        cell = worksheet.cell(row=row_number, column=column)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center")


def _style_section_label(worksheet, row_number: int, max_column: int) -> None:
    """Apply a visible section band without merging calculation cells."""

    from openpyxl.styles import Font, PatternFill

    fill = PatternFill("solid", fgColor="D9EAF7")
    for column in range(1, max_column + 1):
        cell = worksheet.cell(row=row_number, column=column)
        cell.fill = fill
        cell.font = Font(bold=True, color="1F1F1F")


def _format_table_columns(
    worksheet,
    header_row: int,
    data_rows: int,
    number_formats: dict[str, str],
) -> None:
    """Format table columns by their header labels."""

    if data_rows <= 0:
        return
    headers = {
        str(worksheet.cell(row=header_row, column=column).value): column
        for column in range(1, worksheet.max_column + 1)
    }
    for header, number_format in number_formats.items():
        column = headers.get(header)
        if column is None:
            continue
        for row in range(header_row + 1, header_row + data_rows + 1):
            worksheet.cell(row=row, column=column).number_format = number_format


def _autofit_report_sheet(worksheet, max_width: int = 34) -> None:
    """Set readable bounded widths for populated worksheet columns."""

    from openpyxl.styles import Alignment
    from openpyxl.utils import get_column_letter

    for column in range(1, worksheet.max_column + 1):
        values = [worksheet.cell(row=row, column=column).value for row in range(1, worksheet.max_row + 1)]
        width = max((len(str(value)) for value in values if value is not None), default=0) + 2
        worksheet.column_dimensions[get_column_letter(column)].width = min(max(width, 10), max_width)

    for row in worksheet.iter_rows():
        for cell in row:
            if isinstance(cell.value, str) and len(cell.value) > max_width:
                cell.alignment = Alignment(vertical="top", wrap_text=True)


def _write_performance_by_year_sheet(
    writer: pd.ExcelWriter,
    performance_by_year: pd.DataFrame,
) -> None:
    """Write and format the Performance_By_Year analysis sheet."""

    sheet_name = "Performance_By_Year"
    display_frame = performance_by_year.copy()
    if "Is_Partial_Year" in display_frame.columns:
        display_frame["Is_Partial_Year"] = display_frame["Is_Partial_Year"].map(
            {True: "Yes", False: "No"}
        )
    display_frame.to_excel(writer, sheet_name=sheet_name, index=False)
    worksheet = writer.sheets[sheet_name]
    _style_table_header(worksheet, 1, max(1, len(display_frame.columns)))
    _format_table_columns(
        worksheet,
        header_row=1,
        data_rows=len(performance_by_year),
        number_formats={
            "Total_Return": "0.00%;[Red](0.00%);-",
            "Annual_Volatility": "0.00%;[Red](0.00%);-",
            "Sharpe": "0.00",
            "Max_Drawdown": "0.00%;[Red](0.00%);-",
            "Full_CAGR_If_Year_Flat": "0.00%;[Red](0.00%);-",
        },
    )
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions
    _autofit_report_sheet(worksheet)


def _write_return_attribution_sheet(
    writer: pd.ExcelWriter,
    period_attr: pd.DataFrame,
    ticker_attr: pd.DataFrame,
    exclusion_attr: pd.DataFrame,
    methodology: pd.DataFrame,
) -> None:
    """Write the three Return_Attribution tables into one documented sheet."""

    sheet_name = "Return_Attribution"
    methodology_start = 1
    methodology.to_excel(writer, sheet_name=sheet_name, index=False, startrow=methodology_start)

    period_title = methodology_start + len(methodology) + 3
    period_start = period_title + 1
    period_attr.to_excel(writer, sheet_name=sheet_name, index=False, startrow=period_start)

    ticker_title = period_start + len(period_attr) + 3
    ticker_start = ticker_title + 1
    ticker_attr.to_excel(writer, sheet_name=sheet_name, index=False, startrow=ticker_start)

    exclusion_title = ticker_start + len(ticker_attr) + 3
    exclusion_start = exclusion_title + 1
    exclusion_attr.to_excel(writer, sheet_name=sheet_name, index=False, startrow=exclusion_start)

    worksheet = writer.sheets[sheet_name]
    worksheet.cell(row=1, column=1, value="Methodology")
    worksheet.cell(row=period_title + 1, column=1, value="Rebalance_Period_Attribution")
    worksheet.cell(row=ticker_title + 1, column=1, value="Ticker_Attribution")
    worksheet.cell(row=exclusion_title + 1, column=1, value="Ticker_Exclusion_Stress")

    for row in range(methodology_start + 2, methodology_start + len(methodology) + 2):
        if worksheet.cell(row=row, column=1).value == "Recent_Cutoff_Date":
            worksheet.cell(row=row, column=2).number_format = "yyyy-mm-dd"

    section_specs = [
        (1, methodology_start + 1, len(methodology), len(methodology.columns)),
        (period_title + 1, period_start + 1, len(period_attr), len(period_attr.columns)),
        (ticker_title + 1, ticker_start + 1, len(ticker_attr), len(ticker_attr.columns)),
        (exclusion_title + 1, exclusion_start + 1, len(exclusion_attr), len(exclusion_attr.columns)),
    ]
    for section_row, header_row, _, column_count in section_specs:
        _style_section_label(worksheet, section_row, max(1, column_count))
        _style_table_header(worksheet, header_row, max(1, column_count))

    _format_table_columns(
        worksheet,
        header_row=period_start + 1,
        data_rows=len(period_attr),
        number_formats={
            "Rebalance_Date": "yyyy-mm-dd",
            "Next_Rebalance_Date": "yyyy-mm-dd",
            "Period_Return": "0.00%;[Red](0.00%);-",
            "Log_Return_Contribution": "0.00%;[Red](0.00%);-",
            "Net_Log_Contribution_Share": "0.00%;[Red](0.00%);-",
            "Positive_Log_Contribution_Share": "0.00%;[Red](0.00%);-",
        },
    )
    _format_table_columns(
        worksheet,
        header_row=ticker_start + 1,
        data_rows=len(ticker_attr),
        number_formats={
            "Average_Weight": "0.00%;[Red](0.00%);-",
            "Average_Period_Return": "0.00%;[Red](0.00%);-",
            "Best_Period_Return": "0.00%;[Red](0.00%);-",
            "Worst_Period_Return": "0.00%;[Red](0.00%);-",
            "Simple_Return_Contribution": "0.00%;[Red](0.00%);-",
            "Recent_Simple_Return_Contribution": "0.00%;[Red](0.00%);-",
            "Recent_Contribution_Share": "0.00%;[Red](0.00%);-",
        },
    )
    _format_table_columns(
        worksheet,
        header_row=exclusion_start + 1,
        data_rows=len(exclusion_attr),
        number_formats={
            "Full_CAGR": "0.00%;[Red](0.00%);-",
            "CAGR_If_Excluded_To_Cash": "0.00%;[Red](0.00%);-",
            "CAGR_Difference": "0.00%;[Red](0.00%);-",
        },
    )
    worksheet.freeze_panes = f"A{period_start + 2}"
    _autofit_report_sheet(worksheet, max_width=44)
    worksheet.column_dimensions["B"].width = max(worksheet.column_dimensions["B"].width or 0, 54)


def _future_nyse_trading_days(
    start_date: pd.Timestamp,
    end_date: pd.Timestamp | None,
    rebalance_period: int,
) -> list[pd.Timestamp]:
    """Return holding trading days after start_date and before end_date."""

    start = pd.Timestamp(start_date).normalize()
    if pd.isna(start):
        return []

    if end_date is not None and pd.notna(end_date):
        end = pd.Timestamp(end_date).normalize()
    else:
        end = start + pd.Timedelta(days=max(30, int(rebalance_period) * 4))

    if end <= start:
        return []

    cal = pmc.get_calendar("NYSE")
    valid = cal.valid_days(start, end)
    days = [_as_naive_trading_day(x) for x in valid]
    return [d for d in days if d > start and d < end]


def _active_dynamic_buy_rows(current_ops: pd.DataFrame, as_of: pd.Timestamp) -> pd.DataFrame:
    """Select live buy rows eligible for a forward TP/SL schedule."""

    if current_ops is None or current_ops.empty:
        return pd.DataFrame()
    if "Action" not in current_ops.columns:
        return pd.DataFrame()

    ops = current_ops.copy()
    buy_rows = ops[ops["Action"].astype(str).str.lower() == "buy"].copy()
    if buy_rows.empty:
        return pd.DataFrame()

    if "Next_Rebalance_Date" in buy_rows.columns:
        buy_rows["Next_Rebalance_Date"] = pd.to_datetime(
            buy_rows["Next_Rebalance_Date"], errors="coerce"
        )
        buy_rows = buy_rows[
            buy_rows["Next_Rebalance_Date"].isna()
            | (buy_rows["Next_Rebalance_Date"] > as_of)
        ]

    if "Exit_Reason" in buy_rows.columns:
        exit_reason = buy_rows["Exit_Reason"].fillna("").astype(str)
        buy_rows = buy_rows[~exit_reason.isin([EXIT_TAKE_PROFIT, EXIT_STOP_LOSS])]

    return buy_rows


def build_tp_sl_schedule(
    current_ops: pd.DataFrame,
    as_of_date: pd.Timestamp,
    exit_policy: str,
    rebalance_period: int,
    tp_base: float,
    sl_base: float,
    probability: float = 1.0,
) -> pd.DataFrame:
    """Build the forward dynamic TP/SL price schedule for current buys."""

    columns = [
        "Symbol",
        "Rebalance_Date",
        "Schedule_Date",
        "TD",
        "Next_Rebalance_Date",
        "Buy_Price_Close",
        "TP_Return_Threshold",
        "SL_Return_Threshold",
        "TP_Price",
        "SL_Price",
        "Weight",
        "Shares",
    ]

    if exit_policy != EXIT_DYNAMIC_TP_SL:
        return pd.DataFrame(columns=columns)
    if rebalance_period <= 1:
        return pd.DataFrame(columns=columns)

    as_of = pd.Timestamp(as_of_date).normalize()
    active = _active_dynamic_buy_rows(current_ops, as_of)
    if active.empty:
        return pd.DataFrame(columns=columns)

    records = []
    for _, row in active.iterrows():
        symbol = row.get("Symbol")
        rb_date = pd.to_datetime(row.get("Rebalance_Date"), errors="coerce")
        next_rb = pd.to_datetime(row.get("Next_Rebalance_Date"), errors="coerce")
        buy_price = pd.to_numeric(row.get("Buy_Price_Close"), errors="coerce")
        if pd.isna(symbol) or pd.isna(rb_date) or pd.isna(buy_price) or buy_price <= 0:
            continue

        trading_days = _future_nyse_trading_days(rb_date, next_rb, rebalance_period)
        for td, schedule_date in enumerate(trading_days, start=1):
            if td >= rebalance_period:
                break
            tp_ret, sl_ret = thresholds_for_day(
                tp_base,
                sl_base,
                rebalance_period,
                td,
                probability,
            )
            records.append(
                {
                    "Symbol": symbol,
                    "Rebalance_Date": pd.Timestamp(rb_date).normalize(),
                    "Schedule_Date": schedule_date,
                    "TD": td,
                    "Next_Rebalance_Date": (
                        pd.Timestamp(next_rb).normalize() if pd.notna(next_rb) else pd.NaT
                    ),
                    "Buy_Price_Close": float(buy_price),
                    "TP_Return_Threshold": tp_ret,
                    "SL_Return_Threshold": sl_ret,
                    "TP_Price": float(buy_price) * (1.0 + tp_ret),
                    "SL_Price": float(buy_price) * (1.0 - sl_ret),
                    "Weight": row.get("Weight", np.nan),
                    "Shares": row.get("Shares", np.nan),
                }
            )

    return pd.DataFrame(records, columns=columns)


def build_tp_sl_action_checklist(
    schedule_df: pd.DataFrame,
    as_of_date: pd.Timestamp,
    lookahead_dates: int = TP_SL_ACTION_LOOKAHEAD_DATES,
) -> pd.DataFrame:
    """Return the nearest schedule rows for manual alert/check setup."""

    columns = [
        "Date",
        "Symbol",
        "TP_Price",
        "SL_Price",
        "Buy_Price_Close",
        "Weight",
        "Suggested_Check",
    ]
    if schedule_df is None or schedule_df.empty:
        return pd.DataFrame(columns=columns)

    as_of = pd.Timestamp(as_of_date).normalize()
    sched = schedule_df.copy()
    sched["Schedule_Date"] = pd.to_datetime(sched["Schedule_Date"], errors="coerce")
    future = sched[sched["Schedule_Date"] >= as_of].copy()
    if future.empty:
        return pd.DataFrame(columns=columns)

    keep_dates = sorted(future["Schedule_Date"].dropna().unique())[:lookahead_dates]
    checklist = future[future["Schedule_Date"].isin(keep_dates)].copy()
    checklist["Date"] = checklist["Schedule_Date"]
    checklist["Suggested_Check"] = "Set price alert / Check near close"
    return checklist[columns].sort_values(["Date", "Symbol"]).reset_index(drop=True)


def write_rebalance_day_report(
    result: dict,
    status: dict,
    current_ops: pd.DataFrame,
    output_path: str,
    used_live_prices: bool = False,
    mtm_applied: bool = False,
    # 以下参数从调用方传入，避免直接引用模块级配置变量
    strategy_params: Optional[dict] = None,
    selected_factor_indices: Optional[list] = None,
    selected_factor_names: Optional[list[str]] = None,
    composite_factor_sheet: str = "ic_m3_N20",
    strategy_param: str = "",
    rebalance_period: int = 20,
    data_start_offset_days: int = 0,
    rf_rate: float = 0.02,
    price_snapshot_manifest: Optional[dict] = None,
) -> None:
    """
    写入合并后的调仓日报表（单文件，含全部 sheet）。

    Parameters
    ----------
    result : dict
        回测结果（来自 run_detailed_backtest）
    status : dict
        调仓日状态（来自 get_rebalance_day_status）
    current_ops : pd.DataFrame
        当前调仓日操作明细
    output_path : str
        输出 Excel 路径
    used_live_prices : bool
        是否使用了实时价格
    mtm_applied : bool
        是否执行了 MTM 市值重估
    strategy_params : dict
        策略参数字典
    selected_factor_indices : list
        选定因子索引
    selected_factor_names : list[str]
        选定因子名称
    composite_factor_sheet : str
        复合因子方法
    strategy_param : str
        策略参数字符串
    rebalance_period : int
        调仓周期
    data_start_offset_days : int
        数据起始日偏移
    rf_rate : float
        无风险利率
    """
    if strategy_params is None:
        strategy_params = {}
    if selected_factor_indices is None:
        selected_factor_indices = []
    if selected_factor_names is None:
        selected_factor_names = []

    if "error" in result:
        raise ValueError(result["error"])

    params = result.get("params", {})
    as_of = pd.Timestamp(datetime.now().date())
    ret_df = result.get("_ret_df", pd.DataFrame())
    data_coverage_start = (
        str(pd.Timestamp(ret_df.index.min()).date()) if len(ret_df.index) else "-"
    )
    requested_data_start = params.get(
        "requested_data_download_start",
        strategy_params.get("data_download_start_date", ""),
    )
    requested_anchor = params.get(
        "requested_rebalance_anchor",
        strategy_params.get("rebalance_anchor_date", ""),
    )
    effective_start = params.get(
        "effective_rebalance_start",
        params.get("effective_rebalance_anchor", ""),
    )
    price_adjustments = manifest_adjustments_frame(price_snapshot_manifest)
    manifest_status = "available" if price_snapshot_manifest else "not_found"
    manifest_base_file = ""
    manifest_base_run = ""
    manifest_notes = ""
    if price_snapshot_manifest:
        manifest_base_file = str(price_snapshot_manifest.get("base_price_file") or "")
        manifest_base_run = str(price_snapshot_manifest.get("base_run_dir") or "")
        manifest_notes = " | ".join(str(x) for x in price_snapshot_manifest.get("notes") or [])

    daily_returns = result["daily_returns"]
    nav = result["nav"]
    rebalance_returns = result.get("rebalance_returns", pd.Series(dtype=float))

    # 计算绩效指标
    summary = performance_summary(daily_returns, rf=rf_rate, periods_per_year=252)
    total_ret = summary["total_return"]
    ann_ret = summary["annual_return"]
    vol = summary["annual_vol"]
    sharpe = summary["sharpe"]
    max_dd = summary["max_drawdown"]
    max_dd_pct = max_dd * 100 if not np.isnan(max_dd) else float("nan")
    max_loss_duration = summary["max_loss_duration"]
    avg_loss_duration = summary["avg_loss_duration"]
    calmar = summary["calmar"]

    wp_dd, _, _ = worst_period_drawdown(daily_returns, rebalance_returns)
    wp_dd_pct = wp_dd * 100 if not np.isnan(wp_dd) else np.nan
    win_days = int((daily_returns > 0).sum())
    total_days = len(daily_returns)
    win_rate = summary["win_rate"]
    avg_win = float(daily_returns[daily_returns > 0].mean()) if win_days > 0 else 0.0
    loss_days = int((daily_returns < 0).sum())
    avg_loss = float(daily_returns[daily_returns < 0].mean()) if loss_days > 0 else 0.0
    pl_ratio = summary["profit_loss_ratio"]

    def _fmt(v: float, f: str) -> str:
        if isinstance(v, float) and np.isnan(v):
            return "-"
        return f.format(v)

    price_conv_parts = []
    if mtm_applied:
        price_conv_parts.append(
            "未到期持仓：Sell_Price_Close 为 As_Of 日收盘或实时价（假设卖出），见 Sell_Price_Source 列"
        )
    if used_live_prices:
        price_conv_parts.append(
            "调仓日且未收盘：Today_Open=开盘价，Buy_Price_Close=现价（买入估计）"
        )
    if not price_conv_parts:
        price_conv = "Adj Close（收盘价）；T 日收盘执行；未到期持仓已按市值计价列示"
    else:
        price_conv = "；".join(price_conv_parts)

    status_rows = [
        ["Parameter", "Value"],
        ["As_Of_Date", str(as_of.date())],
        ["Is_Rebalance_Today", "是" if status["is_rebalance_today"] else "否"],
        ["Current_Rebalance_Date", str(status["current_rebalance_date"].date()) if status["current_rebalance_date"] else "-"],
        ["Next_Rebalance_Date", str(status["next_rebalance_date"].date()) if status["next_rebalance_date"] else "-"],
        ["Price_Convention", price_conv],
        ["Rebalance_Period_TradingDays", strategy_params.get("rebalance_period", rebalance_period)],
        ["Data_Coverage_Start", data_coverage_start],
        ["Requested_Data_Download_Start", requested_data_start],
        ["Effective_Rebalance_Start", effective_start],
        ["Requested_Rebalance_Anchor", requested_anchor],
        ["Effective_Rebalance_Anchor", effective_start],
        ["Data_Start_Offset_TradingDays", data_start_offset_days],
        ["Preserve_Price_Scale", strategy_params.get("preserve_price_scale", "")],
        ["Price_Scale_Config_Base_Run_Dir", strategy_params.get("price_scale_base_run_dir", "")],
        ["Price_Scale_Manifest", manifest_status],
        ["Price_Scale_Base_Run", manifest_base_run],
        ["Price_Scale_Base_File", manifest_base_file],
        ["Price_Scale_Adjusted_Tickers", len(price_adjustments)],
        ["Price_Scale_Notes", manifest_notes],
        ["---", "---"],
        ["Factor_Indices", str(selected_factor_indices)],
        ["Selected_Factors", ", ".join(selected_factor_names)],
        ["Composite_Factor", composite_factor_sheet],
        ["Composite_Method", _describe_composite_method(composite_factor_sheet)],
        ["Strategy_Param", strategy_param],
        ["Weight_Method", params.get("weight_method", strategy_params.get("weight_method", ""))],
        ["Max_Weight", params.get("max_weight", strategy_params.get("max_weight", ""))],
        ["Group_Num", params.get("group_num", strategy_params.get("group_num", ""))],
        ["Target_Rank", params.get("target_rank", strategy_params.get("target_rank", ""))],
        ["Exit_Policy", params.get("exit_policy", strategy_params.get("exit_policy", ""))],
        ["TP_Base", params.get("tp_base", "")],
        ["SL_Base", params.get("sl_base", "")],
        ["Signal_Probability", params.get("probability", "")],
        ["TP_Count", params.get("tp_count", "")],
        ["SL_Count", params.get("sl_count", "")],
        ["Forced_Close_Count", params.get("forced_close_count", "")],
        ["---", "---"],
        ["Total_Return", _fmt(total_ret, "{:.4f}")],
        ["Annual_Return", _fmt(ann_ret, "{:.4f}")],
        ["Annual_Volatility_Pct", _fmt(vol * 100 if not np.isnan(vol) else float("nan"), "{:.2f}")],
        ["Sharpe_Ratio", _fmt(sharpe, "{:.2f}")],
        ["Max_Drawdown_Pct", _fmt(max_dd_pct, "{:.2f}")],
        ["Max_Loss_Duration_TradingDays", _fmt(max_loss_duration, "{:.0f}")],
        ["Avg_Loss_Duration_TradingDays", _fmt(avg_loss_duration, "{:.2f}")],
        ["Worst_Period_Drawdown_Pct", _fmt(wp_dd_pct, "{:.2f}")],
        ["Calmar_Ratio", _fmt(calmar, "{:.2f}")],
        ["Win_Rate", _fmt(win_rate, "{:.2%}")],
        ["Profit_Loss_Ratio", _fmt(pl_ratio, "{:.2f}")],
    ]

    # 过滤低权重操作
    filtered_ops = filter_weight_lt(current_ops, WEIGHT_FILTER_THRESHOLD, logger=print)
    df_ops_raw = result["operations_df"]
    df_ops_all_filtered = filter_weight_lt(
        df_ops_raw,
        ALL_OPERATIONS_WEIGHT_FILTER_THRESHOLD,
        logger=print,
    )
    df_period = result["period_summary_df"]

    exit_policy = params.get("exit_policy", strategy_params.get("exit_policy", ""))
    performance_by_year = build_performance_by_year(daily_returns, rf_rate=rf_rate)
    period_attr, ticker_attr, exclusion_attr, attribution_methodology = (
        build_return_attribution(
            daily_returns=daily_returns,
            period_summary=df_period,
            operations=df_ops_raw,
            stock_returns=ret_df,
            exit_policy=str(exit_policy),
            rf_rate=rf_rate,
        )
    )

    def _as_float(value, default: float) -> float:
        try:
            if value == "":
                return default
            parsed = float(value)
            if np.isfinite(parsed):
                return parsed
        except (TypeError, ValueError):
            pass
        return default

    tp_base = _as_float(params.get("tp_base", strategy_params.get("tp_base", np.nan)), np.nan)
    sl_base = _as_float(params.get("sl_base", strategy_params.get("sl_base", np.nan)), np.nan)
    probability = _as_float(
        params.get("probability", strategy_params.get("probability", strategy_params.get("tp_sl_probability", 1.0))),
        1.0,
    )
    report_rebalance_period = int(strategy_params.get("rebalance_period", rebalance_period))
    if not np.isfinite(tp_base):
        tp_base = 0.0
    if not np.isfinite(sl_base):
        sl_base = 0.0

    tp_sl_schedule = build_tp_sl_schedule(
        current_ops=filtered_ops,
        as_of_date=as_of,
        exit_policy=str(exit_policy),
        rebalance_period=report_rebalance_period,
        tp_base=tp_base,
        sl_base=sl_base,
        probability=probability,
    )
    tp_sl_checklist = build_tp_sl_action_checklist(tp_sl_schedule, as_of)

    def _nan_to_dash(df: pd.DataFrame) -> pd.DataFrame:
        return df.replace({np.nan: "-"}, inplace=False)

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        pd.DataFrame(status_rows[1:], columns=status_rows[0]).to_excel(
            writer, sheet_name="Rebalance_Config_Status", index=False
        )

        if price_snapshot_manifest:
            if not price_adjustments.empty:
                price_adjustments.to_excel(writer, sheet_name="Price_Scale_Adjustments", index=False)
            else:
                pd.DataFrame(
                    {
                        "Note": ["No price scale adjustments detected"],
                        "Base_Run": [manifest_base_run],
                        "Base_File": [manifest_base_file],
                    }
                ).to_excel(writer, sheet_name="Price_Scale_Adjustments", index=False)
        else:
            pd.DataFrame({"Note": ["No price snapshot manifest found"]}).to_excel(
                writer, sheet_name="Price_Scale_Adjustments", index=False
            )

        if not filtered_ops.empty:
            _nan_to_dash(filtered_ops).to_excel(writer, sheet_name="Current_Operations", index=False)
        else:
            pd.DataFrame({"Note": ["无当前调仓日操作（今日非调仓日或数据不足）"]}).to_excel(
                writer, sheet_name="Current_Operations", index=False
            )

        if str(exit_policy) == EXIT_DYNAMIC_TP_SL:
            if not tp_sl_schedule.empty:
                _nan_to_dash(tp_sl_schedule).to_excel(writer, sheet_name="TP_SL_Schedule", index=False)
            else:
                pd.DataFrame({"Note": ["动态 TP/SL 策略当前无可生成 schedule 的持仓"]}).to_excel(
                    writer, sheet_name="TP_SL_Schedule", index=False
                )

            if not tp_sl_checklist.empty:
                _nan_to_dash(tp_sl_checklist).to_excel(
                    writer, sheet_name="TP_SL_Action_Checklist", index=False
                )
            else:
                pd.DataFrame({"Note": ["当前无今日或近期 TP/SL 人工检查项"]}).to_excel(
                    writer, sheet_name="TP_SL_Action_Checklist", index=False
                )
        else:
            pd.DataFrame({"Note": [f"Exit_Policy={exit_policy}; TP/SL schedule not applicable"]}).to_excel(
                writer, sheet_name="TP_SL_Schedule", index=False
            )
            pd.DataFrame({"Note": [f"Exit_Policy={exit_policy}; TP/SL checklist not applicable"]}).to_excel(
                writer, sheet_name="TP_SL_Action_Checklist", index=False
            )

        future_rb = status.get("future_rebalance_dates", [])
        if future_rb:
            pd.DataFrame({"Future_Rebalance_Date": future_rb}).to_excel(
                writer, sheet_name="Future_Rebalance_Dates", index=False
            )
        else:
            pd.DataFrame({"Note": ["暂无未来调仓日数据"]}).to_excel(
                writer, sheet_name="Future_Rebalance_Dates", index=False
            )

        # All_Operations_All：历史操作明细，过滤 1% 以下的小权重噪音。
        if len(df_ops_all_filtered) > 0:
            _nan_to_dash(df_ops_all_filtered).to_excel(writer, sheet_name="All_Operations_All", index=False)

        if len(df_period) > 0:
            df_period.to_excel(writer, sheet_name="Period_Summary", index=False)

        df_period_2 = build_period_summary_2(df_period)
        df_period_2.to_excel(writer, sheet_name="Period_Summary_2", index=False)

        _write_performance_by_year_sheet(writer, performance_by_year)
        _write_return_attribution_sheet(
            writer,
            period_attr,
            ticker_attr,
            exclusion_attr,
            attribution_methodology,
        )

        df_dr = daily_returns.reset_index()
        df_dr.columns = ["Date", "Daily_Return"]
        df_dr.to_excel(writer, sheet_name="Daily_Returns", index=False)

        df_nav = nav.reset_index()
        df_nav.columns = ["Date", "NAV"]
        df_nav["Cumulative_Return"] = df_nav["NAV"] - 1.0
        df_nav.to_excel(writer, sheet_name="Cumulative_Returns", index=False)

    print(f"调仓日报表已写入: {output_path}")
