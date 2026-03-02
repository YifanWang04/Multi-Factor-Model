"""
策略回测报表生成器 (strategy_report.py)
========================================
输出一份 Excel 工作簿，包含两个 Sheet：

Sheet1 (strategy_statistics)
  - 每行：一个策略参数组合
  - 列：策略名称 + 具体参数 + 全部绩效指标
  - 色阶条件格式：近期收益、年化收益越高越绿；回撤越负越红；夏普越高越绿
  - 嵌入折线图：各资产配置方式在不同分层数量下的年化收益率对比
    （X轴=分层数量，Y轴=年化收益，每条线=一种资产配置方式，按目标组/调仓周期平均）

Sheet2 (strategy_daily_returns)
  - index = 日期，columns = 策略名称，values = 日收益率
  - 附折线图（各策略净值走势，从 Sheet2 导出）
"""

import os
import numpy as np
import pandas as pd
from datetime import datetime

try:
    import openpyxl
    from openpyxl.styles import (
        Font, PatternFill, Alignment, Border, Side, numbers
    )
    from openpyxl.formatting.rule import ColorScaleRule, DataBarRule
    from openpyxl.chart import LineChart, BarChart, Reference
    from openpyxl.chart.series import SeriesLabel
    from openpyxl.utils import get_column_letter
    OPENPYXL_OK = True
except ImportError:
    OPENPYXL_OK = False


# ---------------------------------------------------------------------------
# 列定义
# ---------------------------------------------------------------------------

# 参数列
_PARAM_COLS = [
    ("strategy_name",     "策略名称"),
    ("group_num",         "分层数量"),
    ("target_group",      "目标组号"),
    ("rebalance_period",  "调仓周期(天)"),
    ("weight_method",     "资产配置方式"),
]

# 指标列（中文列头，方便阅读）
_METRIC_COLS = [
    ("ret_1d",             "近1日收益"),
    ("ret_1w",             "近1周收益"),
    ("ret_1m",             "近1月收益"),
    ("ret_3m",             "近3月收益"),
    ("ret_6m",             "近半年收益"),
    ("ret_1y",             "近1年收益"),
    ("ret_last_year",      "上一年整年收益"),
    ("annual_return",      "年化收益率"),
    ("annual_vol",         "年化波动率"),
    ("sharpe",             "夏普比率"),
    ("open_win_rate",      "开仓胜率"),
    ("open_pl_ratio",      "开仓盈亏比"),
    ("annual_open_count",  "年化开仓次数"),
    ("annual_profit_count","年化盈利次数"),
    ("max_drawdown",       "最大回撤"),
    ("calmar",             "Calmar比率"),
    ("max_dd_start",       "最大回撤起始日"),
    ("max_dd_end",         "最大回撤结束日"),
]

_ALL_COLS = _PARAM_COLS + _METRIC_COLS

# 需要百分比格式的指标
_PCT_KEYS = {
    "ret_1d", "ret_1w", "ret_1m", "ret_3m", "ret_6m",
    "ret_1y", "ret_last_year", "annual_return", "annual_vol",
    "open_win_rate", "max_drawdown",
}

# 越大越好（绿）/ 越小越好（红：max_drawdown）
_HIGHER_BETTER = {
    "ret_1d", "ret_1w", "ret_1m", "ret_3m", "ret_6m",
    "ret_1y", "ret_last_year", "annual_return", "sharpe",
    "open_win_rate", "open_pl_ratio", "calmar",
    "annual_open_count", "annual_profit_count",
}
_LOWER_BETTER = {"max_drawdown"}   # 越负越红（用反向色阶）


# ---------------------------------------------------------------------------
# 主报告类
# ---------------------------------------------------------------------------

class StrategyReporter:
    """
    Parameters
    ----------
    results       : {strategy_name: result_dict} — 来自 StrategyBacktester.run_grid()
    all_metrics   : {strategy_name: metrics_dict} — 来自 compute_all_metrics()
    config        : strategy_config 模块
    """

    def __init__(self, results: dict, all_metrics: dict, config):
        self.results = results
        self.metrics = all_metrics
        self.config = config

    def write(self, output_path: str) -> None:
        os.makedirs(os.path.dirname(output_path), exist_ok=True)

        sheet1_df = self._build_sheet1_df()
        sheet2_df = self._build_sheet2_df()

        if OPENPYXL_OK:
            self._write_with_format(output_path, sheet1_df, sheet2_df)
        else:
            with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
                sheet1_df.to_excel(writer, sheet_name="strategy_statistics", index=False)
                sheet2_df.to_excel(writer, sheet_name="strategy_daily_returns")
        print(f"策略回测报表已写入: {output_path}")

    # ------------------------------------------------------------------
    # 构建 DataFrame
    # ------------------------------------------------------------------

    def _build_sheet1_df(self) -> pd.DataFrame:
        """构建统计指标汇总表（一行一个策略）。"""
        rows = []
        for name, m in self.metrics.items():
            row = {"strategy_name": name}
            row["group_num"] = m.get("group_num", np.nan)
            row["target_group"] = m.get("target_group", np.nan)
            row["rebalance_period"] = m.get("rebalance_period", np.nan)
            row["weight_method"] = m.get("weight_method", "")
            for key, _ in _METRIC_COLS:
                row[key] = m.get(key, np.nan)
            rows.append(row)

        df = pd.DataFrame(rows)
        col_order = [c for c, _ in _ALL_COLS]
        df = df.reindex(columns=col_order)

        # 日期列转字符串（Excel 不支持 pd.Timestamp 直接写入时区）
        for date_col in ("max_dd_start", "max_dd_end"):
            df[date_col] = df[date_col].apply(
                lambda x: x.strftime("%Y-%m-%d")
                if isinstance(x, (pd.Timestamp, datetime)) and not pd.isnull(x)
                else x
            )
        return df

    def _build_sheet2_df(self) -> pd.DataFrame:
        """构建日频收益率宽表（行=日期，列=策略名）。"""
        series_dict = {}
        for name, res in self.results.items():
            dr = res.get("daily_returns", pd.Series(dtype=float))
            if len(dr) > 0:
                series_dict[name] = dr

        if not series_dict:
            return pd.DataFrame()

        df = pd.DataFrame(series_dict)
        df.index.name = "Date"
        df.sort_index(inplace=True)
        return df

    # ------------------------------------------------------------------
    # Excel 写入与格式化
    # ------------------------------------------------------------------

    def _write_with_format(
        self, output_path: str, sheet1_df: pd.DataFrame, sheet2_df: pd.DataFrame
    ) -> None:
        with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
            # ── Sheet1 ────────────────────────────────────────────────
            header_row = 1            # 标题行偏移（1-based）
            data_start_row = 2        # 数据行起始（1-based）
            sheet1_df.to_excel(
                writer, sheet_name="strategy_statistics",
                index=False, startrow=header_row - 1,
            )

            ws1 = writer.sheets["strategy_statistics"]
            n_rows = len(sheet1_df)
            n_cols = len(sheet1_df.columns)

            self._format_sheet1_header(ws1, n_cols)
            self._format_sheet1_columns(ws1, sheet1_df, data_start_row, n_rows)
            self._apply_color_scale(ws1, sheet1_df, data_start_row, n_rows)
            self._add_pivot_chart(ws1, sheet1_df, data_start_row, n_rows, n_cols)

            # ── Sheet2 ────────────────────────────────────────────────
            if len(sheet2_df) > 0:
                sheet2_df.to_excel(
                    writer, sheet_name="strategy_daily_returns"
                )
                ws2 = writer.sheets["strategy_daily_returns"]
                self._format_sheet2(ws2, sheet2_df)
                self._add_nav_chart(ws2, sheet2_df)

    # ------------------------------------------------------------------
    # Sheet1 格式
    # ------------------------------------------------------------------

    def _format_sheet1_header(self, ws, n_cols: int) -> None:
        header_fill = PatternFill("solid", fgColor="1F497D")
        header_font = Font(bold=True, color="FFFFFF", size=10)
        for col_idx in range(1, n_cols + 1):
            cell = ws.cell(row=1, column=col_idx)
            # 替换列头为中文名
            key = list(dict(_ALL_COLS).keys())[col_idx - 1] if col_idx - 1 < len(_ALL_COLS) else None
            display = dict(_ALL_COLS).get(key, cell.value) if key else cell.value
            cell.value = display
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.row_dimensions[1].height = 36

    def _format_sheet1_columns(
        self, ws, df: pd.DataFrame, data_start_row: int, n_rows: int
    ) -> None:
        """百分比格式 + 列宽自适应。"""
        pct_fmt = "0.00%"
        num_fmt = "0.00"
        date_fmt = "@"  # 文本

        col_keys = list(df.columns)
        for col_idx, key in enumerate(col_keys, start=1):
            col_letter = get_column_letter(col_idx)

            # 列宽
            max_len = max(
                len(str(df.iloc[r][key]) if not pd.isnull(df.iloc[r][key]) else "")
                for r in range(len(df))
            ) if len(df) > 0 else 0
            header_len = len(dict(_ALL_COLS).get(key, key))
            ws.column_dimensions[col_letter].width = max(
                min(max_len + 2, 22), header_len + 2, 10
            )

            # 数值格式
            for row_idx in range(data_start_row, data_start_row + n_rows):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.alignment = Alignment(horizontal="center", vertical="center")
                if key in _PCT_KEYS:
                    cell.number_format = pct_fmt
                elif key in ("max_dd_start", "max_dd_end", "strategy_name", "weight_method"):
                    cell.number_format = date_fmt
                elif key not in ("group_num", "target_group", "rebalance_period"):
                    cell.number_format = num_fmt

        # 冻结首行
        ws.freeze_panes = "A2"

    def _apply_color_scale(
        self, ws, df: pd.DataFrame, data_start_row: int, n_rows: int
    ) -> None:
        """对关键指标列应用三色色阶条件格式。"""
        if n_rows < 2:
            return
        col_keys = list(df.columns)
        for col_idx, key in enumerate(col_keys, start=1):
            col_letter = get_column_letter(col_idx)
            rng = (
                f"{col_letter}{data_start_row}:"
                f"{col_letter}{data_start_row + n_rows - 1}"
            )
            if key in _HIGHER_BETTER:
                # 低→白，中→浅绿，高→深绿
                rule = ColorScaleRule(
                    start_type="percentile", start_value=10,
                    start_color="FFFFFF",
                    mid_type="percentile", mid_value=50,
                    mid_color="C6EFCE",
                    end_type="percentile", end_value=90,
                    end_color="006100",
                )
                ws.conditional_formatting.add(rng, rule)
            elif key in _LOWER_BETTER:
                # max_drawdown：值越负越差，越接近 0 越好；低（最负）→深红，高（接近0）→白
                rule = ColorScaleRule(
                    start_type="percentile", start_value=10,
                    start_color="9C0006",
                    mid_type="percentile", mid_value=50,
                    mid_color="FFEB9C",
                    end_type="percentile", end_value=90,
                    end_color="FFFFFF",
                )
                ws.conditional_formatting.add(rng, rule)

    # ------------------------------------------------------------------
    # Sheet1 图表（年化收益 × 分层数量 × 资产配置方式）
    # ------------------------------------------------------------------

    def _add_pivot_chart(
        self, ws, df: pd.DataFrame, data_start_row: int, n_rows: int, n_data_cols: int
    ) -> None:
        """
        在 Sheet1 数据区域右侧插入折线图：
          X轴  = 分层数量（GROUP_NUMS）
          Y轴  = 年化收益率（按各配置方式平均，再按调仓周期/目标组平均）
          系列 = 资产配置方式（weight_method）
        """
        if n_rows < 2 or "annual_return" not in df.columns:
            return

        group_nums = sorted(df["group_num"].dropna().unique())
        weight_methods = sorted(df["weight_method"].dropna().unique())

        if len(group_nums) < 2 or not weight_methods:
            return

        # 构造透视汇总表（分层数量 × 资产配置方式 → 平均年化收益）
        pivot = (
            df.groupby(["group_num", "weight_method"])["annual_return"]
            .mean()
            .unstack("weight_method")
            .reindex(index=group_nums)
        )

        # 将透视表写入 Sheet1 右侧的辅助区域（不影响主数据）
        chart_data_start_col = n_data_cols + 2   # 留一空列
        chart_data_start_row = 1

        # 写表头
        ws.cell(row=chart_data_start_row, column=chart_data_start_col, value="分层数量")
        for j, wm in enumerate(pivot.columns):
            ws.cell(
                row=chart_data_start_row, column=chart_data_start_col + 1 + j, value=wm
            )

        # 写数据
        for i, (gn, row_s) in enumerate(pivot.iterrows()):
            r = chart_data_start_row + 1 + i
            ws.cell(row=r, column=chart_data_start_col, value=gn)
            for j, wm in enumerate(pivot.columns):
                val = row_s.get(wm, np.nan)
                ws.cell(row=r, column=chart_data_start_col + 1 + j,
                        value=None if pd.isnull(val) else round(float(val), 6))

        n_pivot_rows = len(pivot)
        n_pivot_cols = len(pivot.columns)

        # 创建折线图
        chart = LineChart()
        chart.title = "年化收益率 × 分层数量（各配置方式）"
        chart.style = 10
        chart.y_axis.title = "年化收益率"
        chart.x_axis.title = "分层数量"
        chart.height = 14
        chart.width = 24

        # Y轴数据（每种资产配置方式一个系列）
        for j in range(n_pivot_cols):
            data_ref = Reference(
                ws,
                min_col=chart_data_start_col + 1 + j,
                min_row=chart_data_start_row,
                max_row=chart_data_start_row + n_pivot_rows,
            )
            chart.add_data(data_ref, titles_from_data=True)

        # X轴标签
        cats = Reference(
            ws,
            min_col=chart_data_start_col,
            min_row=chart_data_start_row + 1,
            max_row=chart_data_start_row + n_pivot_rows,
        )
        chart.set_categories(cats)

        # 图表放置位置：数据区域下方
        anchor_row = data_start_row + n_rows + 3
        anchor_col = 1
        ws.add_chart(chart, f"{get_column_letter(anchor_col)}{anchor_row}")

    # ------------------------------------------------------------------
    # Sheet2 格式 + 净值走势图
    # ------------------------------------------------------------------

    def _format_sheet2(self, ws, df: pd.DataFrame) -> None:
        """Sheet2 基础格式：冻结首行首列，列宽。"""
        ws.freeze_panes = "B2"
        ws.column_dimensions["A"].width = 14
        for col_idx in range(2, len(df.columns) + 2):
            ws.column_dimensions[get_column_letter(col_idx)].width = 14

        # 百分比格式
        pct_fmt = "0.00%"
        for row in ws.iter_rows(
            min_row=2, max_row=ws.max_row, min_col=2, max_col=ws.max_column
        ):
            for cell in row:
                cell.number_format = pct_fmt
                cell.alignment = Alignment(horizontal="right")

    def _add_nav_chart(self, ws, df: pd.DataFrame) -> None:
        """
        在 Sheet2 中插入净值折线图（累计收益 = (1+r).cumprod()-1）。
        策略数量超过 10 时只绘制年化收益前 10 名（避免图例混乱）。
        """
        if len(df) == 0:
            return

        # 计算 NAV 并选取前 10 名策略（按最终 NAV 排序）
        nav_df = (1.0 + df).cumprod()

        max_series = 10
        if nav_df.shape[1] > max_series:
            final_nav = nav_df.iloc[-1].sort_values(ascending=False)
            top_cols = final_nav.index[:max_series].tolist()
            nav_df = nav_df[top_cols]

        # 将 NAV 写入辅助区域（Sheet2 最后列右侧）
        n_data_cols = df.shape[1] + 1  # +1 for Date index col
        nav_start_col = n_data_cols + 2
        nav_start_row = 1

        ws.cell(row=nav_start_row, column=nav_start_col, value="Date_nav")
        for j, col in enumerate(nav_df.columns):
            ws.cell(row=nav_start_row, column=nav_start_col + 1 + j, value=col)

        for i, (dt, row_s) in enumerate(nav_df.iterrows()):
            r = nav_start_row + 1 + i
            ws.cell(row=r, column=nav_start_col,
                    value=dt.strftime("%Y-%m-%d") if hasattr(dt, "strftime") else str(dt))
            for j, col in enumerate(nav_df.columns):
                val = row_s.get(col, np.nan)
                ws.cell(row=r, column=nav_start_col + 1 + j,
                        value=None if pd.isnull(val) else round(float(val), 6))

        n_nav_rows = len(nav_df)
        n_nav_cols = len(nav_df.columns)

        chart = LineChart()
        chart.title = "策略净值走势（年化收益前10）" if nav_df.shape[1] == max_series else "策略净值走势"
        chart.style = 10
        chart.y_axis.title = "净值"
        chart.x_axis.title = "日期"
        chart.height = 16
        chart.width = 28

        for j in range(n_nav_cols):
            data_ref = Reference(
                ws,
                min_col=nav_start_col + 1 + j,
                min_row=nav_start_row,
                max_row=nav_start_row + n_nav_rows,
            )
            chart.add_data(data_ref, titles_from_data=True)

        cats = Reference(
            ws,
            min_col=nav_start_col,
            min_row=nav_start_row + 1,
            max_row=nav_start_row + n_nav_rows,
        )
        chart.set_categories(cats)

        # 放置于 Sheet2 数据下方
        anchor_row = n_nav_rows + nav_start_row + 3
        ws.add_chart(chart, f"A{anchor_row}")
