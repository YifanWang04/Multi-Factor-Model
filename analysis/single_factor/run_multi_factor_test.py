"""
多因子集中测试 (run_multi_factor_test.py)
=====================================
输入：多个因子 Excel、收益率 Excel
处理：对每个因子调用单因子测试流程（单周期），汇总 IC/多空/多头超额等指标
输出：多因子集中测试报表 Excel
  - Sheet1 (factor_test_statistics) : 因子 × 指标汇总表
      列：ic_mean, ic_ir, ic_t_value, rank_ic_mean, rank_ic_ir, rank_ic_t_value,
          group_rank_ic_mean, group_rank_ic_ir, group_rank_ic_t_value,
          long_annual_return, long_sharpe,
          long_excess_annual, long_excess_sharpe,
          ic_p_value, rank_ic_p_value, group_rank_ic_p_value
          （多空 ls_* 与空头 short_* 已注释，可按需启用）
      色阶：p 值越小越绿；其余指标越大越红
  - Sheet2 (factor_cum_ic)      : 日期 × 因子 = 累计 IC，附折线图
  - Sheet3 (factor_LE_cum_ret)  : 日期 × 因子 = 多头累计超额收益率（多头 - 市场），附折线图
  - Sheet4 (factor_L_cum_ret)   : 日期 × 因子 = 多头累计收益率，附折线图
"""
import os
import sys
import numpy as np
import pandas as pd
from datetime import datetime
from types import SimpleNamespace

_SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
_PROJECT_ROOT = os.path.dirname(os.path.dirname(_SCRIPT_DIR))
if _SCRIPT_DIR not in sys.path:
    sys.path.insert(0, _SCRIPT_DIR)
if _PROJECT_ROOT not in sys.path:
    sys.path.insert(0, _PROJECT_ROOT)

from config import SingleFactorConfig
from rebalance_manager import RebalancePeriodManager
from ic import ICAnalyzerEnhanced
from grouping import GrouperEnhanced
from backtest import LongOnlyBacktest, ShortOnlyBacktest
from performance import PerformanceAnalyzer

try:
    import openpyxl
    from openpyxl.chart import LineChart, Reference
    from openpyxl.chart.axis import Scaling
    from openpyxl.formatting.rule import ColorScaleRule
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    Scaling = None


# ---------------------------------------------------------------------------
# 数据加载
# ---------------------------------------------------------------------------

def load_return_data(price_file, return_column="Return"):
    """从价格 Excel 加载日频收益率，与单因子一致。"""
    price_data = pd.read_excel(price_file, sheet_name=None)
    ret = pd.DataFrame()
    for ticker, df in price_data.items():
        if "Date" not in df.columns or "Adj Close" not in df.columns:
            continue
        df = df.copy()
        df["Date"] = pd.to_datetime(df["Date"])
        df.set_index("Date", inplace=True)
        if return_column in df.columns:
            ret[ticker] = df[return_column]
        else:
            ret[ticker] = df["Adj Close"].pct_change()
    ret = ret.replace([np.inf, -np.inf], np.nan)
    return ret


def load_factor(factor_file, sheet_name=0):
    """加载单因子 Excel 的指定 sheet，index 为日期。"""
    factor = pd.read_excel(factor_file, sheet_name=sheet_name, index_col=0)
    factor.index = pd.to_datetime(factor.index)
    factor = factor.apply(pd.to_numeric, errors="coerce")
    return factor


def iter_factors_from_files(factor_files, get_factor_display_name):
    """
    遍历所有因子：每个 Excel 的每个 sheet 作为一个因子。
    产出 (factor_name, factor_df) 供 run_one_factor_one_period 使用。
    """
    for path in factor_files:
        xl = pd.ExcelFile(path)
        base_name = get_factor_display_name(path)
        for sheet in xl.sheet_names:
            try:
                factor = load_factor(path, sheet_name=sheet)
                if factor.empty or len(factor) < 2:
                    continue
                name = f"{base_name}_{sheet}" if len(xl.sheet_names) > 1 else base_name
                yield name, factor
            except Exception:
                continue


# ---------------------------------------------------------------------------
# 单因子单周期运行（复用单因子逻辑，并计算基准与多头超额）
# ---------------------------------------------------------------------------

def _empty_factor_record(config):
    """因子无有效调仓期时返回的占位结果（全 NaN/空序列），保证 Sheet1/2/3/4 不报错。"""
    nan_series = pd.Series(dtype=float)
    empty_nav = pd.Series(dtype=float)
    return {
        "ic_df": pd.DataFrame({"IC": [], "Rank_IC": []}),
        "ic_stats": {
            "IC": {"Mean": np.nan, "IR": np.nan, "t_value": np.nan, "p_value": np.nan},
            "Rank_IC": {"Mean": np.nan, "IR": np.nan, "t_value": np.nan, "p_value": np.nan},
        },
        "group_ic_stats": {
            "Group_IC": {"Mean": np.nan, "IR": np.nan, "t_value": np.nan, "p_value": np.nan},
            "Group_Rank_IC": {"Mean": np.nan, "IR": np.nan, "t_value": np.nan, "p_value": np.nan},
        },
        "ls_stats": {"Annual_Return": np.nan, "Sharpe": np.nan},
        "short_combined_stats": {"Annual_Return": np.nan, "Sharpe": np.nan},
        "long_annual_return": np.nan,
        "long_sharpe": np.nan,
        "long_excess_annual": np.nan,
        "long_excess_sharpe": np.nan,
        "long_excess_returns": nan_series,
        "long_excess_nav": empty_nav,
        "long_nav": empty_nav,
        "ls_returns": nan_series,
        "group_returns": pd.DataFrame(),
        "ret_periods": pd.DataFrame(),
    }


def run_one_factor_one_period(factor, ret, rebalance_period, config):
    """
    对单一因子、单一调仓周期跑完整流程，返回指标与时间序列。
    返回 dict 含：ic_df, ic_stats, group_ic_stats, ls_stats, short_combined_stats,
    long_excess_annual, long_excess_sharpe, long_excess_returns, benchmark_returns,
    ls_returns, group_returns, ret_periods 等。
    若因子与收益日期无重叠或调仓期为 0，返回全 NaN/空序列，避免缺行。
    """
    manager = RebalancePeriodManager(factor, ret, rebalance_period)
    factor_periods, ret_periods = manager.align_factor_return_by_period()

    if len(factor_periods) == 0:
        return _empty_factor_record(config)

    ic_analyzer = ICAnalyzerEnhanced(factor_periods, ret_periods)
    ic_df = ic_analyzer.calculate_ic()
    ic_stats = {
        "IC": ic_analyzer.calculate_statistics(ic_df["IC"]),
        "Rank_IC": ic_analyzer.calculate_statistics(ic_df["Rank_IC"]),
    }

    grouper = GrouperEnhanced(
        factor_periods, config.GROUP_NUM, config.WEIGHT_METHOD
    )
    group_dict = grouper.split()
    weight_dict = grouper.get_group_weights(group_dict)
    group_returns = grouper.calculate_group_returns(
        group_dict, ret_periods, weight_dict
    )

    group_ic_df = ic_analyzer.calculate_group_ic(group_dict, group_returns)
    group_ic_stats = {
        "Group_IC": ic_analyzer.calculate_statistics(group_ic_df["Group_IC"]),
        "Group_Rank_IC": ic_analyzer.calculate_statistics(
            group_ic_df["Group_Rank_IC"]
        ),
    }

    cols = group_returns.columns.tolist()
    if len(cols) < 4:
        index_gr = (
            ret_periods.index
            if len(group_returns) == 0
            else group_returns.index
        )
        group_returns = pd.DataFrame(
            np.zeros((len(index_gr), 4)),
            index=index_gr,
            columns=[1, 2, 3, 4],
        )
        top2_cols = [4, 3]
        bottom2_cols = [1, 2]
    else:
        top2_cols = cols[-2:]
        bottom2_cols = cols[:2]

    long_combined_returns = group_returns[top2_cols].mean(axis=1)
    short_combined_raw = group_returns[bottom2_cols].mean(axis=1)

    ls_returns = (
        long_combined_returns
        - short_combined_raw
        - 2 * config.TRANSACTION_COST
    )
    ls_nav = (1 + ls_returns).cumprod()
    ls_perf = PerformanceAnalyzer(
        ls_nav, ls_returns, config.RISK_FREE_RATE
    )
    ls_stats = ls_perf.calculate_metrics()

    short_combined_returns = (
        -group_returns[bottom2_cols].mean(axis=1)
        - config.TRANSACTION_COST
    )
    short_combined_nav = (1 + short_combined_returns).cumprod()
    short_combined_perf = PerformanceAnalyzer(
        short_combined_nav, short_combined_returns, config.RISK_FREE_RATE
    )
    short_combined_stats = short_combined_perf.calculate_metrics()

    # 基准：全市场等权
    benchmark_returns = ret_periods.mean(axis=1)
    # 多头收益（扣单边成本）
    long_ret = long_combined_returns - config.TRANSACTION_COST
    long_nav = (1 + long_ret).cumprod()
    # 多头绝对收益指标
    long_perf = PerformanceAnalyzer(long_nav, long_ret, config.RISK_FREE_RATE)
    long_metrics = long_perf.calculate_metrics()
    long_annual_return = long_metrics["Annual_Return"]
    long_sharpe = long_metrics["Sharpe"]
    # 多头超额收益 = 多头收益 - 基准收益
    long_excess_returns = long_ret - benchmark_returns
    long_excess_nav = (1 + long_excess_returns).cumprod()
    long_excess_perf = PerformanceAnalyzer(
        long_excess_nav, long_excess_returns, config.RISK_FREE_RATE
    )
    long_excess_metrics = long_excess_perf.calculate_metrics()
    long_excess_annual = long_excess_metrics["Annual_Return"]
    long_excess_sharpe = long_excess_metrics["Sharpe"]

    return {
        "ic_df": ic_df,
        "ic_stats": ic_stats,
        "group_ic_stats": group_ic_stats,
        "ls_stats": ls_stats,
        "short_combined_stats": short_combined_stats,
        "long_annual_return": long_annual_return,
        "long_sharpe": long_sharpe,
        "long_excess_annual": long_excess_annual,
        "long_excess_sharpe": long_excess_sharpe,
        "long_excess_returns": long_excess_returns,
        "long_excess_nav": long_excess_nav,
        "long_nav": long_nav,
        "ls_returns": ls_returns,
        "group_returns": group_returns,
        "ret_periods": ret_periods,
    }


# ---------------------------------------------------------------------------
# 汇总表与写 Excel
# ---------------------------------------------------------------------------

def build_sheet1_df(records, factor_names):
    """
    records: list of dict (each from run_one_factor_one_period)
    factor_names: list of str
    返回 DataFrame: index=因子名, columns=英文指标名（一行一因子）
    """
    rows = []
    for name, rec in zip(factor_names, records):
        ic_s = rec["ic_stats"]["IC"]
        rank_s = rec["ic_stats"]["Rank_IC"]
        gr_s = rec["group_ic_stats"]["Group_Rank_IC"]
        ls_s = rec["ls_stats"]
        sh_s = rec["short_combined_stats"]
        rows.append({
            "ic_mean": ic_s["Mean"],
            "ic_ir": ic_s["IR"],
            "ic_t_value": ic_s["t_value"],
            "rank_ic_mean": rank_s["Mean"],
            "rank_ic_ir": rank_s["IR"],
            "rank_ic_t_value": rank_s["t_value"],
            "group_rank_ic_mean": gr_s["Mean"],
            "group_rank_ic_ir": gr_s["IR"],
            "group_rank_ic_t_value": gr_s["t_value"],
            # "ls_annual_return": ls_s["Annual_Return"],
            # "ls_sharpe": ls_s["Sharpe"],
            "long_annual_return": rec["long_annual_return"],
            "long_sharpe": rec["long_sharpe"],
            "long_excess_annual": rec["long_excess_annual"],
            "long_excess_sharpe": rec["long_excess_sharpe"],
            # "short_annual_return": sh_s["Annual_Return"],
            # "short_sharpe": sh_s["Sharpe"],
            "ic_p_value": ic_s["p_value"],
            "rank_ic_p_value": rank_s["p_value"],
            "group_rank_ic_p_value": gr_s["p_value"],
        })
    df = pd.DataFrame(rows, index=factor_names)
    df.index.name = "factor_name"
    return df


def align_cumulative_series(series_dict, fill=np.nan):
    """series_dict: {factor_name: Series with date index}. 对齐到所有日期的并集。"""
    all_dates = sorted(set().union(*[set(s.index) for s in series_dict.values()]))
    out = pd.DataFrame(index=all_dates)
    for name, s in series_dict.items():
        out[name] = s.reindex(all_dates)
    return out


def build_sheet2_df(records, factor_names):
    """累计 IC：每期 IC 的累计和。"""
    cum_ic = {}
    for name, rec in zip(factor_names, records):
        ic_series = rec["ic_df"]["IC"].copy()
        cum_ic[name] = ic_series.cumsum()
    return align_cumulative_series(cum_ic)


def build_long_excess_df(records, factor_names):
    """Sheet3：多头累计超额收益率（多头收益 - 市场基准），(1+long_excess).cumprod() - 1"""
    cum_excess = {}
    for name, rec in zip(factor_names, records):
        nav = rec["long_excess_nav"]
        cum_excess[name] = nav - 1.0
    return align_cumulative_series(cum_excess)


def build_long_cumret_df(records, factor_names):
    """Sheet4：多头累计收益率，(1+long_ret).cumprod() - 1"""
    cum_long = {}
    for name, rec in zip(factor_names, records):
        nav = rec["long_nav"]
        cum_long[name] = nav - 1.0
    return align_cumulative_series(cum_long)


def write_excel_with_format(out_path, sheet1_df, sheet2_df, sheet3_df, sheet4_df):
    """
    写入 Excel：Sheet1 色阶，Sheet2/3/4 数据+折线图。

    Parameters
    ----------
    sheet3_df : 多头累计超额收益率（多头 - 市场基准）→ Sheet3 factor_LE_cum_ret
    sheet4_df : 多头累计收益率                         → Sheet4 factor_L_cum_ret
    """
    SHEET1_NAME = "factor_test_statistics"
    SHEET2_NAME = "factor_cum_ic"
    SHEET3_NAME = "factor_LE_cum_ret"
    SHEET4_NAME = "factor_L_cum_ret"

    if not OPENPYXL_AVAILABLE:
        engine = "openpyxl"
        try:
            import openpyxl as _  # noqa: F401
        except ImportError:
            engine = "xlsxwriter"
        with pd.ExcelWriter(out_path, engine=engine) as writer:
            sheet1_df.to_excel(writer, sheet_name=SHEET1_NAME)
            sheet2_df.to_excel(writer, sheet_name=SHEET2_NAME)
            sheet3_df.to_excel(writer, sheet_name=SHEET3_NAME)
            sheet4_df.to_excel(writer, sheet_name=SHEET4_NAME)
        return

    def _cell_val(v):
        """统一处理 NaN / None，写入 Excel 单元格前转成 None。"""
        if v is None:
            return None
        try:
            if pd.isna(v):
                return None
        except (TypeError, ValueError):
            pass
        return v

    def _write_timeseries_ws(ws, df):
        """
        将时序 DataFrame 写入 worksheet，保证：
          行 1 = 表头（"date" + 因子名列）
          行 2+ = 数据（日期字符串 + 数值）
        完全不依赖 dataframe_to_rows，避免版本差异产生空行。
        """
        ws.column_dimensions["A"].width = 14
        ws.cell(row=1, column=1, value="date")
        for c_idx, col_name in enumerate(df.columns, 2):
            ws.cell(row=1, column=c_idx, value=col_name)
        for r_idx, (idx_val, row_vals) in enumerate(zip(df.index, df.values), 2):
            if isinstance(idx_val, (pd.Timestamp, datetime)):
                date_str = idx_val.strftime("%Y-%m-%d")
            else:
                date_str = str(idx_val) if idx_val is not None else None
            ws.cell(row=r_idx, column=1, value=date_str)
            for c_idx, val in enumerate(row_vals, 2):
                ws.cell(row=r_idx, column=c_idx, value=_cell_val(val))

    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = SHEET1_NAME

    # Sheet1: 行1 = 表头（index列名 + 各指标列名），行2+ = 数据
    ws1.cell(row=1, column=1, value=sheet1_df.index.name or "factor_name")
    for c_idx, col_name in enumerate(sheet1_df.columns, 2):
        ws1.cell(row=1, column=c_idx, value=col_name)
    for r_idx, (idx_val, row_vals) in enumerate(zip(sheet1_df.index, sheet1_df.values), 2):
        ws1.cell(row=r_idx, column=1, value=idx_val)
        for c_idx, val in enumerate(row_vals, 2):
            ws1.cell(row=r_idx, column=c_idx, value=_cell_val(val))

    # Sheet1: 色阶按列竖向应用，3-color Green-Yellow-Red
    #   p 值：低 = 好 → min=绿, mid=黄, max=红
    #   其余指标：高 = 好 → min=红, mid=黄, max=绿
    # 颜色沿用 Excel 内置 Green-Yellow-Red 色板
    _GREEN  = "63BE7B"
    _YELLOW = "FFEB84"
    _RED    = "F8696B"

    p_value_cols = ["ic_p_value", "rank_ic_p_value", "group_rank_ic_p_value"]
    col_names = list(sheet1_df.columns)
    data_rows = len(sheet1_df) + 1
    for c_idx, col_name in enumerate(col_names, 2):
        cell_range = (
            f"{openpyxl.utils.get_column_letter(c_idx)}2:"
            f"{openpyxl.utils.get_column_letter(c_idx)}{data_rows}"
        )
        if col_name in p_value_cols:
            # p 值：越小越好 → min 绿, max 红
            ws1.conditional_formatting.add(
                cell_range,
                ColorScaleRule(
                    start_type="min",        start_color=_GREEN,
                    mid_type="percentile",   mid_value=50,  mid_color=_YELLOW,
                    end_type="max",          end_color=_RED,
                ),
            )
        else:
            # 一般指标：越大越好 → min 红, max 绿
            ws1.conditional_formatting.add(
                cell_range,
                ColorScaleRule(
                    start_type="min",        start_color=_RED,
                    mid_type="percentile",   mid_value=50,  mid_color=_YELLOW,
                    end_type="max",          end_color=_GREEN,
                ),
            )

    # Sheet2: 行=调仓日，列=因子名，值=累计 IC
    ws2 = wb.create_sheet(SHEET2_NAME)
    _write_timeseries_ws(ws2, sheet2_df)

    n_rows = len(sheet2_df) + 1
    n_cols = len(sheet2_df.columns) + 1
    chart2 = LineChart()
    chart2.title = "Cumulative IC"
    chart2.x_axis.title = "日期"
    chart2.y_axis.title = "累计IC"
    chart2.width = 28
    chart2.height = 16
    _lbl_skip2 = max(1, len(sheet2_df) // 12)
    chart2.x_axis.tickLblSkip = _lbl_skip2
    chart2.x_axis.tickMarkSkip = _lbl_skip2
    if Scaling is not None:
        vals = sheet2_df.values
        valid = vals[np.isfinite(vals)] if vals.size else []
        if len(valid) > 0:
            mn, mx = float(np.nanmin(valid)), float(np.nanmax(valid))
            pad = max((mx - mn) * 0.05, 0.5) if mx != mn else 0.5
            chart2.y_axis.scaling = Scaling(min=mn - pad, max=mx + pad)
            step = (mx - mn + 2 * pad) / 10
            if step > 0:
                chart2.y_axis.majorUnit = max(round(step, 2), 0.1)
    data_ref = Reference(ws2, min_col=2, min_row=1, max_col=n_cols, max_row=n_rows)
    cat_ref = Reference(ws2, min_col=1, min_row=2, max_row=n_rows)
    chart2.add_data(data_ref, titles_from_data=True)
    chart2.set_categories(cat_ref)
    ws2.add_chart(
        chart2,
        openpyxl.utils.get_column_letter(n_cols + 2) + "1",
    )

    # Sheet3: 行=调仓日，列=因子名，值=多头累计超额收益率（多头 - 市场基准）
    ws3 = wb.create_sheet(SHEET3_NAME)
    _write_timeseries_ws(ws3, sheet3_df)

    n_rows3 = len(sheet3_df) + 1
    n_cols3 = len(sheet3_df.columns) + 1
    chart3 = LineChart()
    chart3.title = "Long Cumulative Excess Return (vs Market)"
    chart3.x_axis.title = "日期"
    chart3.y_axis.title = "累计超额收益率"
    chart3.width = 28
    chart3.height = 16
    _lbl_skip3 = max(1, len(sheet3_df) // 12)
    chart3.x_axis.tickLblSkip = _lbl_skip3
    chart3.x_axis.tickMarkSkip = _lbl_skip3
    if Scaling is not None:
        vals3 = sheet3_df.values
        valid3 = vals3[np.isfinite(vals3)] if vals3.size else []
        if len(valid3) > 0:
            mn3, mx3 = float(np.nanmin(valid3)), float(np.nanmax(valid3))
            pad3 = max((mx3 - mn3) * 0.05, 0.02) if mx3 != mn3 else 0.02
            chart3.y_axis.scaling = Scaling(min=mn3 - pad3, max=mx3 + pad3)
            step3 = (mx3 - mn3 + 2 * pad3) / 10
            if step3 > 0:
                chart3.y_axis.majorUnit = max(round(step3, 4), 0.01)
    data_ref3 = Reference(ws3, min_col=2, min_row=1, max_col=n_cols3, max_row=n_rows3)
    cat_ref3 = Reference(ws3, min_col=1, min_row=2, max_row=n_rows3)
    chart3.add_data(data_ref3, titles_from_data=True)
    chart3.set_categories(cat_ref3)
    ws3.add_chart(
        chart3,
        openpyxl.utils.get_column_letter(n_cols3 + 2) + "1",
    )

    # Sheet4: 行=调仓日，列=因子名，值=多头累计收益率
    ws4 = wb.create_sheet(SHEET4_NAME)
    _write_timeseries_ws(ws4, sheet4_df)

    n_rows4 = len(sheet4_df) + 1
    n_cols4 = len(sheet4_df.columns) + 1
    chart4 = LineChart()
    chart4.title = "Long Cumulative Return"
    chart4.x_axis.title = "日期"
    chart4.y_axis.title = "累计收益率"
    chart4.width = 28
    chart4.height = 16
    _lbl_skip4 = max(1, len(sheet4_df) // 12)
    chart4.x_axis.tickLblSkip = _lbl_skip4
    chart4.x_axis.tickMarkSkip = _lbl_skip4
    if Scaling is not None:
        vals4 = sheet4_df.values
        valid4 = vals4[np.isfinite(vals4)] if vals4.size else []
        if len(valid4) > 0:
            mn4, mx4 = float(np.nanmin(valid4)), float(np.nanmax(valid4))
            pad4 = max((mx4 - mn4) * 0.05, 0.02) if mx4 != mn4 else 0.02
            chart4.y_axis.scaling = Scaling(min=mn4 - pad4, max=mx4 + pad4)
            step4 = (mx4 - mn4 + 2 * pad4) / 10
            if step4 > 0:
                chart4.y_axis.majorUnit = max(round(step4, 4), 0.01)
    data_ref4 = Reference(ws4, min_col=2, min_row=1, max_col=n_cols4, max_row=n_rows4)
    cat_ref4 = Reference(ws4, min_col=1, min_row=2, max_row=n_rows4)
    chart4.add_data(data_ref4, titles_from_data=True)
    chart4.set_categories(cat_ref4)
    ws4.add_chart(
        chart4,
        openpyxl.utils.get_column_letter(n_cols4 + 2) + "1",
    )

    wb.save(out_path)


# ---------------------------------------------------------------------------
# 主流程
# ---------------------------------------------------------------------------

def run_multi_factor_test(
    factor_files=None,
    price_file=None,
    rebalance_period=5,
    output_dir=None,
    output_name="multi_factor_test_report.xlsx",
):
    """
    多因子集中测试主入口。
    factor_files: list of str, 因子 Excel 路径
    price_file: str, 价格/收益 Excel
    rebalance_period: int
    output_dir / output_name: 输出目录与文件名
    """
    from multi_factor_config import (
        FACTOR_FILES,
        FACTOR_PROCESSED_DIR,
        PRICE_FILE,
        RETURN_COLUMN,
        OUTPUT_DIR,
        OUTPUT_EXCEL_NAME,
        GROUP_NUM,
        WEIGHT_METHOD,
        RISK_FREE_RATE,
        TRANSACTION_COST,
        get_factor_display_name,
        get_all_factor_files,
    )

    factor_files = factor_files or FACTOR_FILES
    if not factor_files:
        factor_files = get_all_factor_files(FACTOR_PROCESSED_DIR)
    if not factor_files:
        raise FileNotFoundError(
            f"未找到任何因子文件，请检查目录: {FACTOR_PROCESSED_DIR}，"
            "或于 multi_factor_config 中设置 FACTOR_FILES"
        )
    price_file = price_file or PRICE_FILE
    output_dir = output_dir or OUTPUT_DIR
    output_name = output_name or OUTPUT_EXCEL_NAME

    config = SimpleNamespace(
        GROUP_NUM=GROUP_NUM,
        WEIGHT_METHOD=WEIGHT_METHOD,
        RISK_FREE_RATE=RISK_FREE_RATE,
        TRANSACTION_COST=TRANSACTION_COST,
    )

    os.makedirs(output_dir, exist_ok=True)
    ret = load_return_data(price_file, RETURN_COLUMN)
    factor_names = []
    records = []
    factor_list = list(iter_factors_from_files(factor_files, get_factor_display_name))
    if not factor_list:
        raise ValueError(
            "未找到任何有效因子，请检查因子 Excel 是否含多 sheet 且每 sheet 有至少 2 行有效数据"
        )
    for i, (name, factor) in enumerate(factor_list):
        print(f"[{i+1}/{len(factor_list)}] 因子: {name}")
        rec = run_one_factor_one_period(factor, ret, rebalance_period, config)
        if len(rec.get("ic_df", [])) == 0 and len(rec.get("ret_periods", [])) == 0:
            print(f"    警告: 该因子与收益日期无重叠或调仓期为 0，Sheet1/2/3/4 中该因子为缺省或空")
        factor_names.append(name)
        records.append(rec)

    sheet1_df = build_sheet1_df(records, factor_names)
    sheet2_df = build_sheet2_df(records, factor_names)
    sheet3_df = build_long_excess_df(records, factor_names)   # Sheet3: 多头累计超额
    sheet4_df = build_long_cumret_df(records, factor_names)   # Sheet4: 多头累计收益率

    out_path = os.path.join(output_dir, output_name)
    write_excel_with_format(out_path, sheet1_df, sheet2_df, sheet3_df, sheet4_df)
    print(f"报表已写入: {out_path}")
    return out_path


def main():
    from multi_factor_config import (
        FACTOR_FILES,
        FACTOR_PROCESSED_DIR,
        PRICE_FILE,
        OUTPUT_DIR,
        REBALANCE_PERIODS,
        get_all_factor_files,
    )

    factor_list = FACTOR_FILES if FACTOR_FILES else get_all_factor_files(FACTOR_PROCESSED_DIR)
    print("=" * 60)
    print("多因子集中测试（多周期）")
    print("=" * 60)
    print(f"因子数量: {len(factor_list)}")
    print(f"调仓周期: {REBALANCE_PERIODS}")
    print(f"输出目录: {OUTPUT_DIR}")
    for i, p in enumerate(factor_list, 1):
        print(f"  [{i}] {os.path.basename(p)}")

    for period in REBALANCE_PERIODS:
        print(f"\n{'=' * 60}")
        print(f"调仓周期: {period} 天")
        print("=" * 60)
        output_name = f"multi_factor_test_report_P{period}.xlsx"
        run_multi_factor_test(
            factor_files=factor_list,
            price_file=PRICE_FILE,
            rebalance_period=period,
            output_dir=OUTPUT_DIR,
            output_name=output_name,
        )

    print("\n" + "=" * 60)
    print("多因子集中测试完成（全部周期）")
    print("=" * 60)


if __name__ == "__main__":
    main()
