"""
因子共线性分析 (run_collinearity_analysis.py)
=============================================
输入：多个因子 Excel、收益率 Excel
处理：共线性分析

输出：多因子共线性分析报表 Excel
  Sheet1 (factor_corr_matrices)
    Matrix1 (beta_corr)   : 各因子收益率序列之间的 Pearson 相关性矩阵
                            -- 每个调仓期对每个因子做一元 OLS 回归，得到因子收益率 beta
                            -- 对 beta 时序计算相关矩阵
    Matrix2 (factor_corr) : 截面因子值相关性的时序均值矩阵
                            -- 每个调仓日计算 N×N 因子截面相关矩阵（跨标的）
                            -- 对所有调仓日取均值

  Sheet2 (corr_series)
    行=调仓日，列=配对因子名（A vs B），值=当日截面 Pearson 相关系数
    附合并折线图，反映每对因子的截面相关性随时间的变化

  Sheet3 (cum_corr)
    Sheet2 的历史滚动累计和
    斜率陡表示长期相关性高；斜率趋近于零表示两因子统计上不相关
"""
import os
import sys
import numpy as np
import pandas as pd
from datetime import datetime
from itertools import combinations
from types import SimpleNamespace

_SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
_PROJECT_ROOT = os.path.dirname(os.path.dirname(_SCRIPT_DIR))
if _SCRIPT_DIR not in sys.path:
    sys.path.insert(0, _SCRIPT_DIR)
if _PROJECT_ROOT not in sys.path:
    sys.path.insert(0, _PROJECT_ROOT)

from config import SingleFactorConfig
from rebalance_manager import RebalancePeriodManager

try:
    import openpyxl
    from openpyxl.chart import LineChart, Reference
    from openpyxl.chart.axis import Scaling
    from openpyxl.formatting.rule import ColorScaleRule
    from openpyxl.utils.dataframe import dataframe_to_rows
    from openpyxl.styles import Font, PatternFill, Alignment
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    Scaling = None


# ---------------------------------------------------------------------------
# 数据加载（与 run_multi_factor_test 保持一致）
# ---------------------------------------------------------------------------

def load_return_data(price_file, return_column="Return"):
    """从价格 Excel 加载收益率数据，与单因子一致。"""
    price_data = pd.read_excel(price_file, sheet_name=None)
    ret = pd.DataFrame()
    for ticker, df in price_data.items():
        if "Date" not in df.columns or "Adj Close" not in df.columns:
            continue
        df = df.copy()
        df["Date"] = pd.to_datetime(df["Date"])
        df.set_index("Date", inplace=True)
        if return_column in df.columns:
            ret[ticker] = df[return_column]
        else:
            ret[ticker] = df["Adj Close"].pct_change()
    ret = ret.replace([np.inf, -np.inf], np.nan)
    return ret


def load_factor(factor_file, sheet_name=0):
    """加载单因子 Excel 的指定 sheet，index 为日期。"""
    factor = pd.read_excel(factor_file, sheet_name=sheet_name, index_col=0)
    factor.index = pd.to_datetime(factor.index)
    factor = factor.apply(pd.to_numeric, errors="coerce")
    return factor


def iter_factors_from_files(factor_files, get_factor_display_name):
    """
    遍历所有因子文件，产出 (factor_name, factor_df)。
    每个 Excel 的每个 sheet 作为一个独立因子。
    """
    for path in factor_files:
        xl = pd.ExcelFile(path)
        base_name = get_factor_display_name(path)
        for sheet in xl.sheet_names:
            try:
                factor = load_factor(path, sheet_name=sheet)
                if factor.empty or len(factor) < 2:
                    continue
                name = f"{base_name}_{sheet}" if len(xl.sheet_names) > 1 else base_name
                yield name, factor
            except Exception:
                continue


# ---------------------------------------------------------------------------
# 核心分析：调仓期对齐
# ---------------------------------------------------------------------------

def compute_factor_aligned_data(factor_dict, ret, rebalance_period):
    """
    对每个因子运行 RebalancePeriodManager，得到调仓期对齐的因子截面与累计收益。

    Returns
    -------
    aligned : dict {factor_name: (factor_periods_df, ret_periods_df)}
        factor_periods_df : index=调仓日, columns=标的代码, 值=前一期因子值
        ret_periods_df    : index=调仓日, columns=标的代码, 值=期间累计收益率
    """
    aligned = {}
    for name, factor in factor_dict.items():
        manager = RebalancePeriodManager(factor, ret, rebalance_period)
        fp, rp = manager.align_factor_return_by_period()
        if len(fp) > 0:
            aligned[name] = (fp, rp)
    return aligned


# ---------------------------------------------------------------------------
# 核心分析：OLS Beta 估计（Matrix1 原材料）
# ---------------------------------------------------------------------------

def _ols_beta(f_vals, r_vals):
    """
    OLS 一元回归 beta 估计：r = alpha + beta * f。
    使用解析式 beta = Cov(f, r) / Var(f)，避免矩阵求逆的数值不稳定。
    样本量 < 3 或因子方差趋近于零时返回 NaN。
    """
    mask = np.isfinite(f_vals) & np.isfinite(r_vals)
    n = mask.sum()
    if n < 3:
        return np.nan
    f_c = f_vals[mask]
    r_c = r_vals[mask]
    f_demean = f_c - f_c.mean()
    var_f = np.dot(f_demean, f_demean)
    if var_f < 1e-12:
        return np.nan
    return np.dot(f_demean, r_c) / var_f


def compute_beta_series(aligned):
    """
    对每个因子的每个调仓期，通过 OLS 一元回归估计因子收益率（beta）。

    Parameters
    ----------
    aligned : dict {factor_name: (factor_periods, ret_periods)}

    Returns
    -------
    beta_df : DataFrame
        index=调仓日（所有因子的并集）, columns=因子名
        值为该因子在该调仓期的 OLS beta 估计；无法估计时为 NaN
    """
    beta_series_dict = {}
    for name, (fp, rp) in aligned.items():
        betas = []
        dates = []
        for date in fp.index:
            f_row = fp.loc[date].dropna()
            r_row = rp.loc[date].dropna()
            common = f_row.index.intersection(r_row.index)
            beta = _ols_beta(
                f_row.loc[common].values.astype(float),
                r_row.loc[common].values.astype(float),
            ) if len(common) >= 3 else np.nan
            betas.append(beta)
            dates.append(date)
        beta_series_dict[name] = pd.Series(betas, index=dates)

    if not beta_series_dict:
        return pd.DataFrame()

    all_dates = sorted(set().union(*[set(s.index) for s in beta_series_dict.values()]))
    beta_df = pd.DataFrame(
        {name: s.reindex(all_dates) for name, s in beta_series_dict.items()}
    )
    return beta_df


# ---------------------------------------------------------------------------
# 核心分析：截面因子相关系数（Matrix2 / Sheet2 原材料）
# ---------------------------------------------------------------------------

def _cross_section_corr(series_a, series_b):
    """
    计算两个 Series（跨标的截面）的 Pearson 相关系数。
    取共同非 NaN 标的；有效样本 < 3 或常量输入返回 NaN。
    """
    common = series_a.index.intersection(series_b.index)
    if len(common) < 3:
        return np.nan
    a = series_a.loc[common].values.astype(float)
    b = series_b.loc[common].values.astype(float)
    mask = np.isfinite(a) & np.isfinite(b)
    if mask.sum() < 3:
        return np.nan
    a_c, b_c = a[mask], b[mask]
    if np.std(a_c) < 1e-12 or np.std(b_c) < 1e-12:
        return np.nan
    return float(np.corrcoef(a_c, b_c)[0, 1])


def compute_cross_sectional_corr_series(aligned):
    """
    对每个调仓日，计算所有配对因子的截面 Pearson 相关系数（跨标的）。

    Parameters
    ----------
    aligned : dict {factor_name: (factor_periods, ret_periods)}

    Returns
    -------
    corr_series : DataFrame
        index=调仓日, columns="factorA vs factorB"
        值=当日截面 Pearson 相关系数；当日任一因子无数据时为 NaN
    """
    names = list(aligned.keys())
    pairs = list(combinations(names, 2))
    if not pairs:
        return pd.DataFrame()

    all_dates = sorted(set().union(*[set(fp.index) for fp, _ in aligned.values()]))
    pair_cols = [f"{a} vs {b}" for a, b in pairs]
    records = {col: [] for col in pair_cols}

    for date in all_dates:
        slices = {
            name: fp.loc[date].dropna()
            for name, (fp, _) in aligned.items()
            if date in fp.index
        }
        for (a, b), col in zip(pairs, pair_cols):
            if a in slices and b in slices:
                records[col].append(_cross_section_corr(slices[a], slices[b]))
            else:
                records[col].append(np.nan)

    return pd.DataFrame(records, index=all_dates)


# ---------------------------------------------------------------------------
# 构建 Matrix1 与 Matrix2
# ---------------------------------------------------------------------------

def build_matrix1_beta_corr(beta_df):
    """
    Matrix1：因子收益率序列的 Pearson 相关性矩阵。
    对 beta_df 中全为 NaN 的行先删除，再调用 .corr()。
    """
    if beta_df.empty or beta_df.shape[1] < 2:
        return pd.DataFrame()
    return beta_df.dropna(how="all").corr()


def build_matrix2_factor_corr(aligned):
    """
    Matrix2：截面因子相关性的时序均值矩阵。

    每个调仓日：
      1. 取所有因子在该日的截面值，计算 N×N Pearson 相关矩阵（跨标的）
      2. 若某因子在该日无数据，该因子行/列填 NaN

    对所有调仓日的相关矩阵取 nanmean，得到 Matrix2。
    """
    names = list(aligned.keys())
    n = len(names)
    if n < 2:
        return pd.DataFrame()

    all_dates = sorted(set().union(*[set(fp.index) for fp, _ in aligned.values()]))
    corr_stack = []

    for date in all_dates:
        slices = {
            name: fp.loc[date].dropna()
            for name, (fp, _) in aligned.items()
            if date in fp.index
        }
        if len(slices) < 2:
            continue

        # 取当前日有数据的因子列表，计算局部相关矩阵
        available = [nm for nm in names if nm in slices]
        common_stocks = slices[available[0]].index
        for nm in available[1:]:
            common_stocks = common_stocks.intersection(slices[nm].index)
        if len(common_stocks) < 3:
            continue

        factor_matrix = pd.DataFrame(
            {nm: slices[nm].loc[common_stocks].values for nm in available},
            index=common_stocks,
            columns=available,
        ).astype(float).dropna()
        if len(factor_matrix) < 3:
            continue

        # 局部 N'×N' 相关矩阵，扩展到全 N×N（缺失因子填 NaN）
        local_corr = factor_matrix.corr()
        full_corr = pd.DataFrame(np.nan, index=names, columns=names)
        full_corr.loc[available, available] = local_corr.values
        corr_stack.append(full_corr.values)

    if not corr_stack:
        return pd.DataFrame(np.nan, index=names, columns=names)

    mean_corr = np.nanmean(np.stack(corr_stack, axis=0), axis=0)
    return pd.DataFrame(mean_corr, index=names, columns=names)


# ---------------------------------------------------------------------------
# Excel 写入工具
# ---------------------------------------------------------------------------

def _format_date_value(value):
    """统一转换日期/NaN 以写入 Excel 单元格。"""
    if value is None:
        return None
    if isinstance(value, (pd.Timestamp, datetime)):
        return value.strftime("%Y-%m-%d")
    try:
        if pd.isna(value):
            return None
    except (TypeError, ValueError):
        pass
    return value


def _write_matrix_block(ws, matrix_df, title, start_row, start_col=1):
    """
    将相关性矩阵写入 worksheet 的指定起始位置。
    布局：标题行 → 表头行（列名）→ 数据行（含行名）→ 3-色阶条件格式
    返回：下一个可写的 start_row（留 2 行间距）
    """
    if matrix_df.empty:
        ws.cell(row=start_row, column=start_col, value=title + "（数据不足，无法计算）")
        return start_row + 3

    if OPENPYXL_AVAILABLE:
        title_cell = ws.cell(row=start_row, column=start_col, value=title)
        title_cell.font = Font(bold=True, size=11)
        title_cell.fill = PatternFill("solid", fgColor="D9E1F2")
    else:
        ws.cell(row=start_row, column=start_col, value=title)

    # 表头（因子名）
    for j, col_name in enumerate(matrix_df.columns, start_col + 1):
        cell = ws.cell(row=start_row + 1, column=j, value=str(col_name))
        if OPENPYXL_AVAILABLE:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")

    # 行名 + 数据
    data_start_row = start_row + 2
    for i, (row_name, row_data) in enumerate(matrix_df.iterrows(), data_start_row):
        name_cell = ws.cell(row=i, column=start_col, value=str(row_name))
        if OPENPYXL_AVAILABLE:
            name_cell.font = Font(bold=True)
        for j, val in enumerate(row_data, start_col + 1):
            numeric_val = round(float(val), 6) if pd.notna(val) else None
            data_cell = ws.cell(row=i, column=j, value=numeric_val)
            if OPENPYXL_AVAILABLE:
                data_cell.alignment = Alignment(horizontal="center")

    data_end_row = data_start_row + len(matrix_df) - 1
    n_data_cols = len(matrix_df.columns)

    # 3-色阶条件格式：-1(红) → 0(白) → +1(绿)
    if OPENPYXL_AVAILABLE and len(matrix_df) > 0:
        for col_offset in range(n_data_cols):
            c_idx = start_col + 1 + col_offset
            col_letter = openpyxl.utils.get_column_letter(c_idx)
            cell_range = f"{col_letter}{data_start_row}:{col_letter}{data_end_row}"
            ws.conditional_formatting.add(
                cell_range,
                ColorScaleRule(
                    start_type="num", start_value=-1, start_color="FF4444",
                    mid_type="num",   mid_value=0,   mid_color="FFFFFF",
                    end_type="num",   end_value=1,   end_color="44BB44",
                ),
            )

    return data_end_row + 3  # 留 2 行间距


def _write_time_series_sheet(ws, df, chart_title, y_axis_title, chart_width=30, chart_height=18):
    """
    将时序 DataFrame 写入 worksheet，并附折线图。
    - index 列宽自动扩展
    - NaN 单元格写入 None（Excel 不显示断点）
    """
    ws.column_dimensions["A"].width = 14

    for r_idx, row in enumerate(dataframe_to_rows(df, index=True, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            if r_idx == 1 and c_idx == 1:
                ws.cell(row=r_idx, column=c_idx, value=value)
                continue
            val = _format_date_value(value) if c_idx == 1 else (
                None if (value is not None and pd.isna(value)) else value
            )
            ws.cell(row=r_idx, column=c_idx, value=val)

    if df.empty or not OPENPYXL_AVAILABLE:
        return

    n_rows = len(df) + 1
    n_cols = len(df.columns) + 1

    chart = LineChart()
    chart.title = chart_title
    chart.x_axis.title = "调仓日"
    chart.y_axis.title = y_axis_title
    chart.width = chart_width
    chart.height = chart_height

    if Scaling is not None:
        vals = df.values.astype(float)
        valid = vals[np.isfinite(vals)] if vals.size else np.array([])
        if len(valid) > 0:
            mn, mx = float(np.nanmin(valid)), float(np.nanmax(valid))
            pad = max((mx - mn) * 0.08, 0.05) if mx != mn else 0.05
            chart.y_axis.scaling = Scaling(min=mn - pad, max=mx + pad)

    data_ref = Reference(ws, min_col=2, min_row=1, max_col=n_cols, max_row=n_rows)
    cat_ref = Reference(ws, min_col=1, min_row=2, max_row=n_rows)
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cat_ref)
    ws.add_chart(chart, openpyxl.utils.get_column_letter(n_cols + 2) + "1")


# ---------------------------------------------------------------------------
# 汇总写入
# ---------------------------------------------------------------------------

def write_collinearity_excel(out_path, matrix1, matrix2, corr_series, cum_corr):
    """
    将共线性分析结果写入 Excel 报表。

    Sheet 说明
    ----------
    factor_corr_matrices : Matrix1（beta_corr）+ Matrix2（factor_corr）
    corr_series          : 每个调仓日各配对因子的截面相关系数时序 + 折线图
    cum_corr             : 累计相关系数（历史滚动累计和）+ 折线图
    """
    SHEET1_NAME = "factor_corr_matrices"
    SHEET2_NAME = "corr_series"
    SHEET3_NAME = "cum_corr"

    # fallback：无 openpyxl 格式化时直接写 DataFrame
    if not OPENPYXL_AVAILABLE:
        with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
            if not matrix1.empty:
                matrix1.to_excel(writer, sheet_name=SHEET1_NAME, startrow=0)
            if not matrix2.empty:
                start = (len(matrix1) + 3) if not matrix1.empty else 0
                matrix2.to_excel(writer, sheet_name=SHEET1_NAME, startrow=start)
            corr_series.to_excel(writer, sheet_name=SHEET2_NAME)
            cum_corr.to_excel(writer, sheet_name=SHEET3_NAME)
        return

    wb = openpyxl.Workbook()

    # ---- Sheet1: 两个矩阵 ----
    ws1 = wb.active
    ws1.title = SHEET1_NAME
    ws1.column_dimensions["A"].width = 24

    next_row = _write_matrix_block(
        ws1, matrix1,
        "Matrix1: 因子收益率相关性矩阵 (beta_corr)  "
        "-- 各因子 OLS beta 时序的 Pearson 相关系数",
        start_row=1,
    )
    _write_matrix_block(
        ws1, matrix2,
        "Matrix2: 截面因子相关性均值矩阵 (factor_corr)  "
        "-- 每调仓日跨标的因子相关矩阵的时序均值",
        start_row=next_row,
    )

    # ---- Sheet2: corr_series ----
    ws2 = wb.create_sheet(SHEET2_NAME)
    _write_time_series_sheet(
        ws2, corr_series,
        "Cross-Sectional Factor Correlation Series",
        "截面相关系数",
    )

    # ---- Sheet3: cum_corr ----
    ws3 = wb.create_sheet(SHEET3_NAME)
    _write_time_series_sheet(
        ws3, cum_corr,
        "Cumulative Correlation  (斜率越陡 = 长期相关性越高；斜率趋零 = 因子统计独立)",
        "累计相关系数",
    )

    wb.save(out_path)


# ---------------------------------------------------------------------------
# 主流程
# ---------------------------------------------------------------------------

def run_collinearity_analysis(
    factor_files=None,
    price_file=None,
    rebalance_period=5,
    output_dir=None,
    output_name=None,
):
    """
    因子共线性分析主入口。

    Parameters
    ----------
    factor_files    : list[str] | None  因子 Excel 路径列表；None 时从配置目录自动扫描
    price_file      : str | None        价格/收益 Excel
    rebalance_period: int               调仓周期（天）
    output_dir      : str | None        输出目录
    output_name     : str | None        输出文件名；None 时自动加 period 后缀
    """
    from multi_factor_config import (
        FACTOR_FILES,
        FACTOR_PROCESSED_DIR,
        PRICE_FILE,
        RETURN_COLUMN,
        OUTPUT_DIR,
        get_factor_display_name,
        get_all_factor_files,
    )

    factor_files = factor_files or FACTOR_FILES
    if not factor_files:
        factor_files = get_all_factor_files(FACTOR_PROCESSED_DIR)
    if not factor_files:
        raise FileNotFoundError(
            f"未找到任何因子文件，请检查目录: {FACTOR_PROCESSED_DIR}，"
            "或于 multi_factor_config 中设置 FACTOR_FILES"
        )

    price_file = price_file or PRICE_FILE
    output_dir = output_dir or OUTPUT_DIR
    output_name = output_name or f"factor_collinearity_report_P{rebalance_period}.xlsx"

    os.makedirs(output_dir, exist_ok=True)

    # ---- 加载数据 ----
    print("加载收益率数据...")
    ret = load_return_data(price_file, RETURN_COLUMN)

    factor_dict = {}
    factor_list = list(iter_factors_from_files(factor_files, get_factor_display_name))
    if not factor_list:
        raise ValueError("未找到任何有效因子数据，请检查因子 Excel 文件。")
    for name, factor in factor_list:
        factor_dict[name] = factor

    print(f"因子数量  : {len(factor_dict)}")
    print(f"调仓周期  : {rebalance_period} 天")
    print(f"输出目录  : {output_dir}")

    # ---- 调仓期对齐 ----
    print("对齐调仓期数据...")
    aligned = compute_factor_aligned_data(factor_dict, ret, rebalance_period)
    if len(aligned) < 2:
        raise ValueError(
            f"有效因子不足 2 个（当前 {len(aligned)} 个），无法进行因子对相关性分析。"
        )
    print(f"有效因子数（含收益率对齐）: {len(aligned)}")

    # ---- 计算 beta 序列 ----
    print("计算因子收益率 (OLS beta)...")
    beta_df = compute_beta_series(aligned)
    if not beta_df.empty:
        valid_betas = beta_df.notna().sum()
        print(f"  各因子有效 beta 期数: {valid_betas.to_dict()}")

    # ---- 计算截面相关系数序列 ----
    print("计算截面因子相关系数...")
    corr_series = compute_cross_sectional_corr_series(aligned)
    if not corr_series.empty:
        valid_corr = corr_series.notna().sum()
        print(f"  各因子对有效期数: {valid_corr.to_dict()}")

    # ---- 构建矩阵 ----
    print("构建 Matrix1 (beta_corr) 与 Matrix2 (factor_corr)...")
    matrix1 = build_matrix1_beta_corr(beta_df)
    matrix2 = build_matrix2_factor_corr(aligned)

    # ---- 累计相关系数 ----
    cum_corr = corr_series.cumsum() if not corr_series.empty else pd.DataFrame()

    # ---- 写入 Excel ----
    out_path = os.path.join(output_dir, output_name)
    print(f"写入报表: {out_path}")
    write_collinearity_excel(out_path, matrix1, matrix2, corr_series, cum_corr)
    print(f"报表已写入: {out_path}")
    return out_path


def main():
    from multi_factor_config import (
        FACTOR_FILES,
        FACTOR_PROCESSED_DIR,
        OUTPUT_DIR,
        get_all_factor_files,
    )

    REBALANCE_PERIODS = [1, 5, 10]

    factor_list = FACTOR_FILES if FACTOR_FILES else get_all_factor_files(FACTOR_PROCESSED_DIR)

    print("=" * 60)
    print("因子共线性分析（多调仓周期）")
    print("=" * 60)
    print(f"因子文件数: {len(factor_list)}")
    print(f"调仓周期  : {REBALANCE_PERIODS}")
    print(f"输出目录  : {OUTPUT_DIR}")
    for i, p in enumerate(factor_list, 1):
        print(f"  [{i}] {os.path.basename(p)}")

    for period in REBALANCE_PERIODS:
        print(f"\n{'=' * 60}")
        print(f"调仓周期: {period} 天")
        print("=" * 60)
        run_collinearity_analysis(
            factor_files=factor_list,
            rebalance_period=period,
            output_dir=OUTPUT_DIR,
        )

    print("\n" + "=" * 60)
    print("因子共线性分析完成（全部调仓周期）")
    print("=" * 60)


if __name__ == "__main__":
    main()
