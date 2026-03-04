"""
Walk-Forward Results Analyzer

分析和汇总walk-forward验证结果，输出单一 consolidated Excel 报表：
1. Summary        - 总体概览、配置摘要、最优策略
2. Parameter_Stability - 参数稳定性（含完整绩效指标 + 绿黄红条件格式）
3. Parameter_Sensitivity - 四表敏感性（原始模式：param → avg_sharpe, sharpe_std, count）
4. Walk_Comparison   - 各 Walk 表现对比
5. daily_returns    - 日收益率（行=日期，列=策略）
6. cumulative_returns - 累计收益率
7. 可视化图表
"""

import os
import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
import seaborn as sns
from typing import List, Dict
from datetime import datetime

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.formatting.rule import ColorScaleRule
    from openpyxl.utils import get_column_letter
    OPENPYXL_OK = True
except ImportError:
    OPENPYXL_OK = False


# 参数稳定性报告列定义（与 strategy_report 一致，含更多指标）
_PARAM_COLS = [
    ("strategy_name", "策略名称"),
    ("group_num", "分层数量"),
    ("target_rank", "目标组排名"),
    ("rebalance_period", "调仓周期(天)"),
    ("weight_method", "资产配置方式"),
]
_METRIC_COLS = [
    ("avg_sharpe", "平均夏普"),
    ("sharpe_std", "夏普标准差"),
    ("win_rate", "胜率(Sharpe>0)"),
    ("avg_annual_return", "平均年化收益"),
    ("avg_annual_vol", "平均年化波动"),
    ("avg_max_drawdown", "平均最大回撤"),
    ("avg_calmar", "平均Calmar"),
    ("avg_open_win_rate", "平均开仓胜率"),
    ("avg_open_pl_ratio", "平均开仓盈亏比"),
    ("avg_annual_open_count", "年化开仓次数"),
    ("num_walks", "Walk数"),
    ("consistency_score", "一致性得分"),
]
# 区间收益列（近1月、近3月、近半年、近1年、上一年整年）
_PERIOD_RETURN_COLS = [
    ("ret_1m", "近1月收益"),
    ("ret_3m", "近3月收益"),
    ("ret_6m", "近半年收益"),
    ("ret_1y", "近1年收益"),
    ("ret_prev_year", "上一年整年收益"),
]
_PCT_KEYS = {"win_rate", "avg_annual_return", "avg_annual_vol", "avg_max_drawdown", "avg_open_win_rate",
             "ret_1m", "ret_3m", "ret_6m", "ret_1y", "ret_prev_year"}
_HIGHER_BETTER = {"avg_sharpe", "win_rate", "avg_annual_return", "avg_calmar", "avg_open_win_rate", "avg_open_pl_ratio", "avg_annual_open_count", "consistency_score",
                  "ret_1m", "ret_3m", "ret_6m", "ret_1y", "ret_prev_year"}
_LOWER_BETTER = {"sharpe_std", "avg_max_drawdown"}


class WalkForwardAnalyzer:
    """
    Walk-Forward结果分析器

    输入：所有walk的所有策略结果
    输出：单一 consolidated Excel 报表 + 可视化
    """

    def __init__(self, results: List[Dict], config):
        self.results = results
        self.config = config
        self.results_df = pd.DataFrame(results)

        self.num_walks = self.results_df['walk_id'].nunique()
        self.num_strategies = len(self.results_df.groupby(
            ['group_num', 'target_rank', 'rebalance_period', 'weight_method']
        ))

        os.makedirs(config.OUTPUT_DIR, exist_ok=True)
        self.output_dir = config.OUTPUT_DIR

        # 用于可视化的缓存
        self.stability_df = None
        self.walk_df = None
        self.daily_returns_df = None
        self.cumulative_returns_df = None
        self.output_file = os.path.join(self.output_dir, 'walk_forward_report.xlsx')

    def generate_all_reports(self):
        """生成 consolidated 报告和可视化"""
        print("\nGenerating consolidated walk-forward report...")

        # 1. 构建参数稳定性数据
        print("  1/4: Parameter stability data...")
        self.stability_df = self._build_parameter_stability_df()

        # 2. 构建 sensitivity 四表（原始模式）
        print("  2/5: Parameter sensitivity tables...")
        sensitivity_tables = self._build_sensitivity_tables()

        # 3. 构建 walk 对比数据
        print("  3/5: Walk comparison data...")
        self.walk_df = self._build_walk_comparison_df()

        # 4. 构建 daily / cumulative returns
        print("  4/5: Daily & cumulative returns...")
        self.daily_returns_df = self._build_daily_returns_df()
        self.cumulative_returns_df = self._build_cumulative_returns_df()

        # 4a. 为 parameter_stability 添加区间收益列
        self._add_period_returns_to_stability()

        # 4b. 按夏普从高到低重排 daily/cumulative returns 的列顺序
        self._reorder_returns_by_sharpe()

        # 5. 写入单一 Excel 文件
        print("  5/5: Writing consolidated Excel...")
        self._write_consolidated_excel(sensitivity_tables)

        if self.config.GENERATE_PLOTS:
            print("  Generating visualizations...")
            self.generate_visualizations()

        print(f"\n[OK] Report saved to: {self.output_file}")

    def _build_parameter_stability_df(self) -> pd.DataFrame:
        """按参数组合聚合，计算稳定性指标及 Strategy_backtest 类指标"""
        grouped = self.results_df.groupby(
            ['group_num', 'target_rank', 'rebalance_period', 'weight_method']
        )

        rows = []
        for params, group in grouped:
            gn, tr, rp, wm = params
            def _safe(col):
                return group[col].dropna() if col in group.columns else pd.Series()

            sharpe_vals = _safe('sharpe')
            ann_ret_vals = _safe('annual_return')
            ann_vol_vals = _safe('annual_vol')
            mdd_vals = _safe('max_drawdown')
            calmar_vals = _safe('calmar')
            open_wr = _safe('open_win_rate')
            open_pl = _safe('open_pl_ratio')
            open_cnt = _safe('annual_open_count')

            row = {
                'strategy_name': f"{wm}_{gn}G_Top{tr}_P{rp}d",
                'group_num': gn, 'target_rank': tr,
                'rebalance_period': rp, 'weight_method': wm,
                'avg_sharpe': sharpe_vals.mean(),
                'sharpe_std': sharpe_vals.std() if len(sharpe_vals) > 1 else 0,
                'win_rate': (sharpe_vals > 0).mean(),
                'avg_annual_return': ann_ret_vals.mean(),
                'avg_annual_vol': ann_vol_vals.mean(),
                'avg_max_drawdown': mdd_vals.mean(),
                'avg_calmar': calmar_vals.mean(),
                'avg_open_win_rate': open_wr.mean() if len(open_wr) > 0 else np.nan,
                'avg_open_pl_ratio': open_pl.mean() if len(open_pl) > 0 else np.nan,
                'avg_annual_open_count': open_cnt.mean() if len(open_cnt) > 0 else np.nan,
                'num_walks': len(group),
                'consistency_score': self._compute_consistency_score(sharpe_vals),
            }
            rows.append(row)

        df = pd.DataFrame(rows).sort_values('consistency_score', ascending=False)
        return df

    def _build_sensitivity_tables(self) -> Dict[str, pd.DataFrame]:
        """构建四组参数敏感性表格（原始模式：param → avg_sharpe, sharpe_std, count）"""
        df = self.results_df
        if df['sharpe'].isna().all():
            return {}

        g1 = df.groupby('group_num')['sharpe'].agg(['mean', 'std', 'count']).reset_index()
        g1.columns = ['group_num', 'avg_sharpe', 'sharpe_std', 'count']

        g2 = df.groupby('target_rank')['sharpe'].agg(['mean', 'std', 'count']).reset_index()
        g2.columns = ['target_rank', 'avg_sharpe', 'sharpe_std', 'count']

        g3 = df.groupby('rebalance_period')['sharpe'].agg(['mean', 'std', 'count']).reset_index()
        g3.columns = ['rebalance_period', 'avg_sharpe', 'sharpe_std', 'count']

        g4 = df.groupby('weight_method')['sharpe'].agg(['mean', 'std', 'count']).reset_index()
        g4.columns = ['weight_method', 'avg_sharpe', 'sharpe_std', 'count']

        return {
            'group_num': g1,
            'target_rank': g2,
            'rebalance_period': g3,
            'weight_method': g4,
        }

    def _build_walk_comparison_df(self) -> pd.DataFrame:
        """构建 Walk 对比数据"""
        rows = []
        for walk_id in range(self.num_walks):
            wd = self.results_df[self.results_df['walk_id'] == walk_id]
            if len(wd) == 0:
                continue
            sharpe_vals = wd['sharpe'].dropna()
            if len(sharpe_vals) == 0:
                print(f"  [WARN] Walk {walk_id}: All Sharpe NaN, skipping...")
                continue

            best_idx = wd['sharpe'].idxmax()
            worst_idx = wd['sharpe'].idxmin()
            best_s, worst_s = wd.loc[best_idx], wd.loc[worst_idx]

            train_period = wd.iloc[0]['train_period']
            test_period = wd.iloc[0]['test_period']

            rows.append({
                'walk_id': walk_id,
                'train_start': train_period[0],
                'train_end': train_period[1],
                'test_start': test_period[0],
                'test_end': test_period[1],
                'best_strategy': best_s['strategy_name'],
                'best_sharpe': best_s['sharpe'],
                'worst_strategy': worst_s['strategy_name'],
                'worst_sharpe': worst_s['sharpe'],
                'avg_sharpe': wd['sharpe'].mean(),
                'median_sharpe': wd['sharpe'].median(),
                'sharpe_range': best_s['sharpe'] - worst_s['sharpe'],
                'positive_sharpe_ratio': (wd['sharpe'] > 0).mean(),
            })

        return pd.DataFrame(rows)

    def _build_daily_returns_df(self) -> pd.DataFrame:
        """构建日收益率宽表（行=日期，列=策略名），拼接各 walk 的 daily_returns"""
        series_dict = {}
        for strategy_name in self.results_df['strategy_name'].unique():
            parts = []
            for r in self.results:
                if r.get('strategy_name') != strategy_name:
                    continue
                dr = r.get('daily_returns')
                if dr is not None and hasattr(dr, '__len__') and len(dr) > 0:
                    parts.append(dr)
            if parts:
                concat = pd.concat(parts).sort_index()
                concat = concat[~concat.index.duplicated(keep='last')]
                series_dict[strategy_name] = concat

        if not series_dict:
            return pd.DataFrame()

        df = pd.DataFrame(series_dict)
        df.index.name = "Date"
        df.sort_index(inplace=True)
        return df

    def _build_cumulative_returns_df(self, daily_df: pd.DataFrame = None) -> pd.DataFrame:
        """由日收益率构建累计收益率 (1+r).cumprod()-1"""
        dr = daily_df if daily_df is not None else self.daily_returns_df
        if dr is None or len(dr) == 0:
            return pd.DataFrame()
        cum_ret = (1.0 + dr).cumprod() - 1.0
        cum_ret.index.name = "Date"
        return cum_ret

    def _add_period_returns_to_stability(self):
        """为 parameter_stability 添加近1月/3月/半年/1年/上一年整年收益"""
        if self.stability_df is None or self.daily_returns_df is None or len(self.daily_returns_df) == 0:
            return
        # 交易日近似：1月≈21，3月≈63，半年≈126，1年≈252
        windows = {"ret_1m": 21, "ret_3m": 63, "ret_6m": 126, "ret_1y": 252}
        for col, days in windows.items():
            self.stability_df[col] = np.nan
        self.stability_df["ret_prev_year"] = np.nan

        for idx, row in self.stability_df.iterrows():
            name = row["strategy_name"]
            if name not in self.daily_returns_df.columns:
                continue
            sr = self.daily_returns_df[name].dropna().sort_index()
            if len(sr) == 0:
                continue
            end_dt = sr.index[-1]
            # 近1月、近3月、近半年、近1年
            for col, days in windows.items():
                lookback = sr.iloc[-days:] if len(sr) >= days else sr
                if len(lookback) > 0:
                    self.stability_df.at[idx, col] = float((1.0 + lookback).prod() - 1.0)
            # 上一年整年收益
            prev_year = end_dt.year - 1
            mask = (sr.index.year == prev_year)
            prev_yr_sr = sr[mask]
            if len(prev_yr_sr) > 0:
                self.stability_df.at[idx, "ret_prev_year"] = float((1.0 + prev_yr_sr).prod() - 1.0)

    def _reorder_returns_by_sharpe(self):
        """按 parameter_stability 中策略的夏普从高到低重排 daily_returns 和 cumulative_returns 的列"""
        if (self.stability_df is None or len(self.stability_df) == 0 or
                self.daily_returns_df is None or len(self.daily_returns_df) == 0):
            return
        # 按 avg_sharpe 降序，缺失夏普的排最后
        ordered = self.stability_df.sort_values(
            "avg_sharpe", ascending=False, na_position="last"
        )["strategy_name"].tolist()
        # 仅保留在 returns 中存在的策略
        valid_cols = [c for c in ordered if c in self.daily_returns_df.columns]
        if not valid_cols:
            return
        self.daily_returns_df = self.daily_returns_df[valid_cols]
        if self.cumulative_returns_df is not None and len(self.cumulative_returns_df) > 0:
            self.cumulative_returns_df = self.cumulative_returns_df[[c for c in valid_cols if c in self.cumulative_returns_df.columns]]

    def _write_consolidated_excel(self, sensitivity_tables: Dict[str, pd.DataFrame]):
        """写入单一 Excel 文件"""
        output_path = self.output_file

        if not OPENPYXL_OK:
            with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                self.stability_df.to_excel(writer, sheet_name='Parameter_Stability', index=False)
                self.walk_df.to_excel(writer, sheet_name='Walk_Comparison', index=False)
                for name, t in sensitivity_tables.items():
                    if t is not None and not t.empty:
                        t.to_excel(writer, sheet_name=name[:31], index=False)
                if self.daily_returns_df is not None and len(self.daily_returns_df) > 0:
                    self.daily_returns_df.to_excel(writer, sheet_name='daily_returns')
                if self.cumulative_returns_df is not None and len(self.cumulative_returns_df) > 0:
                    self.cumulative_returns_df.to_excel(writer, sheet_name='cumulative_returns')
            return

        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            self._write_summary_sheet(writer)
            self._write_parameter_stability_sheet(writer)
            self._write_sensitivity_sheet(writer, sensitivity_tables)
            self._write_walk_comparison_sheet(writer)
            self._write_daily_returns_sheet(writer)
            self._write_cumulative_returns_sheet(writer)

            # 移除默认空 sheet（若有）
            for sname in list(writer.book.sheetnames):
                if sname in ('Sheet', 'Sheet1'):
                    writer.book.remove(writer.book[sname])
                    break

    def _write_summary_sheet(self, writer):
        """写入 Summary 概览"""
        ws = writer.book.create_sheet("Summary", 0)
        cfg = self.config

        # 标题
        ws['A1'] = "Walk-Forward Validation Summary"
        ws['A1'].font = Font(bold=True, size=14)
        row = 3

        # 1. 总体统计
        top = self.stability_df.iloc[0] if len(self.stability_df) > 0 else None
        ws[f'A{row}'] = "Overall"
        ws[f'A{row}'].font = Font(bold=True)
        row += 1
        ws[f'A{row}'] = "Total Walks"; ws[f'B{row}'] = self.num_walks; row += 1
        ws[f'A{row}'] = "Total Strategy Combinations"; ws[f'B{row}'] = self.num_strategies; row += 1
        ws[f'A{row}'] = "Total Backtests"; ws[f'B{row}'] = len(self.results_df); row += 1
        row += 2

        # 2. 最优策略
        ws[f'A{row}'] = "Top Strategy (by consistency_score)"
        ws[f'A{row}'].font = Font(bold=True)
        row += 1
        if top is not None:
            ws[f'A{row}'] = "Strategy"; ws[f'B{row}'] = top['strategy_name']; row += 1
            ws[f'A{row}'] = "Avg Sharpe"; ws[f'B{row}'] = round(top['avg_sharpe'], 4); row += 1
            ws[f'A{row}'] = "Sharpe Std"; ws[f'B{row}'] = round(top['sharpe_std'], 4); row += 1
            ws[f'A{row}'] = "Win Rate"; ws[f'B{row}'] = f"{top['win_rate']:.1%}"; row += 1
            ws[f'A{row}'] = "Avg Annual Return"; ws[f'B{row}'] = f"{top['avg_annual_return']:.2%}"; row += 1
            ws[f'A{row}'] = "Avg Max Drawdown"; ws[f'B{row}'] = f"{top['avg_max_drawdown']:.2%}"; row += 1
        row += 2

        # 3. 配置摘要
        ws[f'A{row}'] = "Configuration"
        ws[f'A{row}'].font = Font(bold=True)
        row += 1
        ws[f'A{row}'] = "Training Window"; ws[f'B{row}'] = f"{cfg.TRAINING_WINDOW} days"; row += 1
        ws[f'A{row}'] = "Testing Window"; ws[f'B{row}'] = f"{cfg.TESTING_WINDOW} days"; row += 1
        ws[f'A{row}'] = "Roll Forward Step"; ws[f'B{row}'] = f"{cfg.ROLL_FORWARD_STEP} days"; row += 1
        ws[f'A{row}'] = "Composite Method"; ws[f'B{row}'] = cfg.COMPOSITE_METHOD; row += 1
        ws[f'A{row}'] = "Group Numbers"; ws[f'B{row}'] = str(cfg.GROUP_NUMS); row += 1
        ws[f'A{row}'] = "Target Ranks"; ws[f'B{row}'] = str(cfg.TARGET_GROUP_RANKS); row += 1
        ws[f'A{row}'] = "Rebalance Periods"; ws[f'B{row}'] = str(cfg.REBALANCE_PERIODS); row += 1
        ws[f'A{row}'] = "Weight Methods"; ws[f'B{row}'] = str(cfg.WEIGHT_METHODS); row += 1

        ws.column_dimensions['A'].width = 24
        ws.column_dimensions['B'].width = 32

    def _write_parameter_stability_sheet(self, writer):
        """写入 Parameter_Stability，含绿黄红条件格式"""
        df = self.stability_df
        df.to_excel(writer, sheet_name="Parameter_Stability", index=False, startrow=0)

        ws = writer.sheets["Parameter_Stability"]
        n_rows = len(df)
        n_cols = len(df.columns)
        data_start = 2
        col_keys = list(df.columns)

        # 表头格式
        header_fill = PatternFill("solid", fgColor="1F497D")
        header_font = Font(bold=True, color="FFFFFF", size=10)
        col_name_map = dict(_PARAM_COLS + _METRIC_COLS + _PERIOD_RETURN_COLS)
        for c in range(1, n_cols + 1):
            cell = ws.cell(row=1, column=c)
            key = col_keys[c - 1]
            cell.value = col_name_map.get(key, key)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        ws.freeze_panes = "A2"

        # 数值格式与条件格式
        for col_idx, key in enumerate(col_keys, start=1):
            col_letter = get_column_letter(col_idx)
            rng = f"{col_letter}{data_start}:{col_letter}{data_start + n_rows - 1}"

            for r in range(data_start, data_start + n_rows):
                c = ws.cell(row=r, column=col_idx)
                c.alignment = Alignment(horizontal="center", vertical="center")
                if key in _PCT_KEYS:
                    c.number_format = "0.00%"
                elif key not in ("strategy_name", "weight_method"):
                    c.number_format = "0.00"

            # 三色阶：红(差) → 黄(中) → 绿(好)，与 strategy_backtest_report 一致
            if n_rows >= 2:
                if key in _HIGHER_BETTER:
                    ws.conditional_formatting.add(rng, ColorScaleRule(
                        start_type="percentile", start_value=10, start_color="FF6B6B",  # 红
                        mid_type="percentile", mid_value=50, mid_color="FFEB9C",       # 黄
                        end_type="percentile", end_value=90, end_color="006100",        # 绿
                    ))
                elif key in _LOWER_BETTER:
                    ws.conditional_formatting.add(rng, ColorScaleRule(
                        start_type="percentile", start_value=10, start_color="9C0006",   # 深红(最差)
                        mid_type="percentile", mid_value=50, mid_color="FFEB9C",        # 黄
                        end_type="percentile", end_value=90, end_color="C6EFCE",        # 浅绿(最好)
                    ))

    def _write_sensitivity_sheet(self, writer, tables: Dict[str, pd.DataFrame]):
        """写入 Parameter_Sensitivity，四表排列于同一 sheet（原始模式：param → avg_sharpe, sharpe_std, count）"""
        ws = writer.book.create_sheet("Parameter_Sensitivity")
        ws.column_dimensions["A"].width = 18

        titles = [
            ("group_num", "1. Group Number Sensitivity"),
            ("target_rank", "2. Target Rank Sensitivity"),
            ("rebalance_period", "3. Rebalance Period Sensitivity"),
            ("weight_method", "4. Weight Method Sensitivity"),
        ]

        next_row = 1
        for key, title in titles:
            t = tables.get(key, pd.DataFrame())
            next_row = self._write_sensitivity_table_block(ws, t, title, next_row)

    def _write_sensitivity_table_block(self, ws, df: pd.DataFrame, title: str, start_row: int) -> int:
        """写入单组敏感性表格（表格式），返回下一可用行"""
        if df is None or df.empty:
            ws.cell(row=start_row, column=1, value=title + "（数据不足）")
            return start_row + 3

        title_cell = ws.cell(row=start_row, column=1, value=title)
        title_cell.font = Font(bold=True, size=11)
        title_cell.fill = PatternFill("solid", fgColor="D9E1F2")

        # 表头
        for j, col in enumerate(df.columns, 1):
            cell = ws.cell(row=start_row + 1, column=j, value=str(col))
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")

        # 数据
        for i, (_, r) in enumerate(df.iterrows(), start_row + 2):
            for j, col in enumerate(df.columns, 1):
                v = r[col]
                if pd.notna(v) and isinstance(v, (int, float)) and col != df.columns[0]:
                    cell = ws.cell(row=i, column=j, value=round(float(v), 4))
                else:
                    cell = ws.cell(row=i, column=j, value=v)
                cell.alignment = Alignment(horizontal="center")

        data_end = start_row + 1 + len(df)
        # 对 avg_sharpe 列应用色阶（第2列）
        avg_col = None
        for j, c in enumerate(df.columns):
            if c == 'avg_sharpe':
                avg_col = j + 1
                break
        if avg_col and len(df) >= 2:
            col_letter = get_column_letter(avg_col)
            ws.conditional_formatting.add(
                f"{col_letter}{start_row + 2}:{col_letter}{data_end}",
                ColorScaleRule(start_type="num", start_value=-1, start_color="FF4444",
                              mid_type="num", mid_value=0, mid_color="FFFF00",
                              end_type="num", end_value=2, end_color="44BB44")
            )
        return data_end + 3

    def _write_daily_returns_sheet(self, writer):
        """写入 daily_returns sheet（行=日期，列=策略）"""
        if self.daily_returns_df is None or len(self.daily_returns_df) == 0:
            return
        self.daily_returns_df.to_excel(writer, sheet_name="daily_returns")
        ws = writer.sheets["daily_returns"]
        ws.freeze_panes = "B2"
        ws.column_dimensions["A"].width = 14
        for c in range(2, self.daily_returns_df.shape[1] + 2):
            ws.column_dimensions[get_column_letter(c)].width = 12
        pct_fmt = "0.00%"
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=2, max_col=ws.max_column):
            for cell in row:
                cell.number_format = pct_fmt
                cell.alignment = Alignment(horizontal="right")

    def _write_cumulative_returns_sheet(self, writer):
        """写入 cumulative_returns sheet"""
        if self.cumulative_returns_df is None or len(self.cumulative_returns_df) == 0:
            return
        self.cumulative_returns_df.to_excel(writer, sheet_name="cumulative_returns")
        ws = writer.sheets["cumulative_returns"]
        ws.freeze_panes = "B2"
        ws.column_dimensions["A"].width = 14
        for c in range(2, self.cumulative_returns_df.shape[1] + 2):
            ws.column_dimensions[get_column_letter(c)].width = 12
        pct_fmt = "0.00%"
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=2, max_col=ws.max_column):
            for cell in row:
                cell.number_format = pct_fmt
                cell.alignment = Alignment(horizontal="right")

    def _write_walk_comparison_sheet(self, writer):
        """写入 Walk_Comparison"""
        df = self.walk_df.copy()
        # 日期列转字符串
        for col in ('train_start', 'train_end', 'test_start', 'test_end'):
            if col in df.columns:
                df[col] = df[col].apply(lambda x: x.strftime("%Y-%m-%d") if hasattr(x, 'strftime') else x)

        df.to_excel(writer, sheet_name="Walk_Comparison", index=False)
        ws = writer.sheets["Walk_Comparison"]
        ws.freeze_panes = "A2"
        ws.column_dimensions["A"].width = 12

    def generate_visualizations(self):
        """生成可视化图表（使用内存中的 stability_df / walk_df）"""
        viz_dir = os.path.join(self.output_dir, 'visualizations')
        os.makedirs(viz_dir, exist_ok=True)

        self._plot_sharpe_heatmap(viz_dir)
        self._plot_stability_scatter(viz_dir)
        self._plot_parameter_boxplots(viz_dir)
        self._plot_walk_performance(viz_dir)

    def _plot_sharpe_heatmap(self, output_dir):
        if self.stability_df is None or self.stability_df['avg_sharpe'].isna().all():
            return
        pivot = self.stability_df.pivot_table(
            values='avg_sharpe', index='weight_method', columns='rebalance_period', aggfunc='mean'
        )
        if pivot.empty:
            return
        plt.figure(figsize=(10, 6))
        sns.heatmap(pivot, annot=True, fmt='.3f', cmap='RdYlGn', center=0)
        plt.title('Average Sharpe Ratio by Weight Method and Rebalance Period')
        plt.tight_layout()
        plt.savefig(os.path.join(output_dir, 'sharpe_heatmap.png'), dpi=300)
        plt.close()

    def _plot_stability_scatter(self, output_dir):
        if self.stability_df is None or self.stability_df['avg_sharpe'].isna().all():
            return
        plt.figure(figsize=(10, 6))
        scatter = plt.scatter(
            self.stability_df['avg_sharpe'],
            self.stability_df['sharpe_std'],
            c=self.stability_df['win_rate'], s=100, alpha=0.6, cmap='RdYlGn'
        )
        plt.colorbar(scatter, label='Win Rate')
        plt.xlabel('Average Sharpe Ratio')
        plt.ylabel('Sharpe Std')
        plt.title('Strategy Stability: Return vs. Consistency')
        plt.axvline(x=0, color='r', linestyle='--', alpha=0.3)
        plt.grid(True, alpha=0.3)
        plt.tight_layout()
        plt.savefig(os.path.join(output_dir, 'stability_scatter.png'), dpi=300)
        plt.close()

    def _plot_parameter_boxplots(self, output_dir):
        if self.results_df['sharpe'].isna().all():
            return
        fig, axes = plt.subplots(2, 2, figsize=(14, 10))
        self.results_df.boxplot(column='sharpe', by='group_num', ax=axes[0, 0])
        axes[0, 0].set_title('Sharpe by Group Number')
        self.results_df.boxplot(column='sharpe', by='target_rank', ax=axes[0, 1])
        axes[0, 1].set_title('Sharpe by Target Rank')
        self.results_df.boxplot(column='sharpe', by='rebalance_period', ax=axes[1, 0])
        axes[1, 0].set_title('Sharpe by Rebalance Period')
        self.results_df.boxplot(column='sharpe', by='weight_method', ax=axes[1, 1])
        axes[1, 1].set_title('Sharpe by Weight Method')
        axes[1, 1].tick_params(axis='x', rotation=45)
        plt.suptitle('Parameter Sensitivity Analysis')
        plt.tight_layout()
        plt.savefig(os.path.join(output_dir, 'parameter_boxplots.png'), dpi=300)
        plt.close()

    def _plot_walk_performance(self, output_dir):
        if self.walk_df is None or len(self.walk_df) == 0:
            return
        fig, axes = plt.subplots(2, 1, figsize=(12, 8))
        axes[0].bar(self.walk_df['walk_id'], self.walk_df['avg_sharpe'], alpha=0.7)
        axes[0].axhline(y=0, color='r', linestyle='--', alpha=0.5)
        axes[0].set_ylabel('Average Sharpe')
        axes[0].set_title('Average Sharpe Across Walks')
        axes[0].grid(True, alpha=0.3)
        axes[1].bar(self.walk_df['walk_id'], self.walk_df['sharpe_range'], alpha=0.7, color='orange')
        axes[1].set_ylabel('Sharpe Range (Best - Worst)')
        axes[1].set_title('Strategy Dispersion Across Walks')
        axes[1].grid(True, alpha=0.3)
        plt.tight_layout()
        plt.savefig(os.path.join(output_dir, 'walk_performance.png'), dpi=300)
        plt.close()

    def get_most_robust_strategy(self) -> Dict:
        """获取最稳健策略（基于 consistency_score）"""
        if self.stability_df is None or len(self.stability_df) == 0:
            return {}
        top = self.stability_df.iloc[0]
        return {
            'params': top['strategy_name'],
            'avg_sharpe': top['avg_sharpe'],
            'sharpe_std': top['sharpe_std'],
            'win_rate': top['win_rate'],
            'avg_return': top['avg_annual_return'],
            'avg_mdd': top['avg_max_drawdown'],
            'consistency_score': top['consistency_score']
        }

    def _compute_consistency_score(self, sharpe_values: pd.Series) -> float:
        if len(sharpe_values) == 0:
            return 0.0
        avg = sharpe_values.mean()
        std = sharpe_values.std()
        if pd.isna(std) or std == 0:
            std = 0
        wr = (sharpe_values > 0).mean()
        return avg * 0.4 + (1 / (1 + std)) * 0.3 + wr * 0.3


if __name__ == "__main__":
    print("Walk-Forward Analyzer - requires results from walk_forward_engine.py")
