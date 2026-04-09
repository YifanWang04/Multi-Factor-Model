"""
backfill_close.py
=================
用途：读取最新 output/rebalance_day_*/data/us_top100_daily_2023_present.xlsx，
      对目标日期有 Open/High/Low 但无 Close 的 sheet，
      用 yfinance fast_info.last_price 补全收盘价，写回原文件。

用法（项目根目录）：
  python backfill_close.py
"""

from __future__ import annotations

import sys
import os
import time
import glob

import pandas as pd
import yfinance as yf

# ── 路径注册 ────────────────────────────────────────────────────────────────
_HERE = os.path.dirname(os.path.abspath(__file__))
_DATA_DIR = os.path.join(_HERE, "data")
sys.path.insert(0, _HERE)

from data.data_config import PRICE_FILE, YFINANCE_TICKERS


# ── 目标日期（默认昨天；若想手动指定，改为具体日期） ────────────────────────
# 美东 4/9 凌晨，对应美股交易日 4/8
import datetime
_TODAY = datetime.date.today()
_YESTERDAY = _TODAY - datetime.timedelta(days=1)
TARGET_DATE = pd.Timestamp(_YESTERDAY)   # 可改为 pd.Timestamp("2025-04-08")
print(f"目标补全日期：{TARGET_DATE.date()}")


# ── fetch_live_prices（复用 run_rebalance_day.py 的逻辑） ────────────────────

LIVE_PRICE_MAX_RETRIES = 3
LIVE_PRICE_RETRY_DELAY_BASE = 0.5
LIVE_PRICE_RETRY_DELAY_MULT = 2.0


def fetch_close_prices(symbols: list[str]) -> dict[str, float]:
    """
    通过 yfinance fast_info.last_price 获取收盘价。
    返回 {symbol: close_price}，失败的标的会被跳过。
    """
    result: dict[str, float] = {}
    delay = LIVE_PRICE_RETRY_DELAY_BASE

    for sym in symbols:
        success = False
        last_error = ""
        for attempt in range(LIVE_PRICE_MAX_RETRIES):
            try:
                ticker = yf.Ticker(sym)
                fi = ticker.fast_info
                last_p = getattr(fi, "last_price", None)
                if last_p is not None:
                    result[sym] = float(last_p)
                    success = True
                    break
            except Exception as e:
                last_error = str(e)
            if attempt < LIVE_PRICE_MAX_RETRIES - 1:
                time.sleep(delay)
                delay *= LIVE_PRICE_RETRY_DELAY_MULT

        if not success:
            print(f"  ⚠️ 获取 {sym} 收盘价失败: {last_error}")

    return result


# ── 主逻辑 ─────────────────────────────────────────────────────────────────

def find_latest_output_price_file() -> str | None:
    """查找最新的 output/rebalance_day_*/data/ 目录下的价格文件。"""
    pattern = os.path.join(_HERE, "output", "rebalance_day_*", "data", "us_top100_daily_2023_present.xlsx")
    files = sorted(glob.glob(pattern), key=os.path.getmtime, reverse=True)
    return files[0] if files else None


def main():
    # 优先从最新 output 目录读取价格文件
    price_file = find_latest_output_price_file()
    if price_file:
        print(f"使用最新 output 价格文件: {price_file}")
    else:
        price_file = PRICE_FILE
        print(f"使用默认价格文件: {price_file}")

    if not os.path.isfile(price_file):
        print(f"价格文件不存在: {price_file}")
        sys.exit(1)

    print(f"读取价格文件: {price_file}")
    all_sheets = pd.read_excel(price_file, sheet_name=None)

    # 找出需要补全 Close 的标的
    need_fetch: list[str] = []
    sheet_names_missing: list[str] = []

    for ticker, df in all_sheets.items():
        if "Date" not in df.columns:
            continue
        df_dt = pd.to_datetime(df["Date"], errors="coerce")
        target_rows = df.loc[df_dt == TARGET_DATE.normalize()]
        if target_rows.empty:
            continue
        # 检查 Close 是否缺失
        if "Close" not in df.columns or "Adj Close" not in df.columns:
            continue
        close_vals = target_rows[["Close", "Adj Close"]].values.flatten()
        if all((pd.isna(v) or v == 0) for v in close_vals):
            # 有 Open/High/Low（至少有一列非空）
            ohl_cols = [c for c in ["Open", "High", "Low"] if c in df.columns]
            if ohl_cols:
                ohl_vals = target_rows[ohl_cols].values.flatten()
                if not all(pd.isna(v) or v == 0 for v in ohl_vals):
                    need_fetch.append(ticker)
                    sheet_names_missing.append(ticker)

    if not need_fetch:
        print("没有需要补全 Close 的标的，退出。")
        sys.exit(0)

    print(f"需要补全收盘价的标的 ({len(need_fetch)} 只): {need_fetch}")

    # 拉取收盘价
    print("正在通过 yfinance 获取收盘价...")
    close_prices = fetch_close_prices(need_fetch)

    if not close_prices:
        print("无法获取任何收盘价，退出。")
        sys.exit(1)

    print(f"成功获取 {len(close_prices)} 只收盘价: {close_prices}")

    # 写入 Excel（逐 sheet 更新目标行）
    import openpyxl
    from openpyxl.utils import get_column_letter

    wb = openpyxl.load_workbook(price_file)

    for ticker, close_val in close_prices.items():
        if ticker not in wb.sheetnames:
            print(f"  ⚠️ {ticker} sheet 不存在，跳过")
            continue

        ws = wb[ticker]
        # 找到目标日期的行（第3行开始，数据行；第1行=0，第2行=1，第1行数据=2）
        target_date_str = TARGET_DATE.strftime("%Y-%m-%d")
        target_row_idx = None

        for row in ws.iter_rows(min_row=2):
            cell_date = row[0].value  # A 列是 Date
            if cell_date is None:
                continue
            # 支持 datetime / date / str
            if hasattr(cell_date, "date"):
                cell_date_val = pd.Timestamp(cell_date).normalize()
            elif isinstance(cell_date, str):
                try:
                    cell_date_val = pd.Timestamp(cell_date).normalize()
                except Exception:
                    continue
            else:
                continue

            if cell_date_val == TARGET_DATE.normalize():
                target_row_idx = row[0].row
                break

        if target_row_idx is None:
            print(f"  ⚠️ {ticker} 在 {TARGET_DATE.date()} 无数据行，跳过")
            continue

        # 写入 Close 和 Adj Close
        close_col_idx = None
        adj_close_col_idx = None
        for col_idx, cell in enumerate(ws[target_row_idx], start=1):
            if cell.value == "Close":
                close_col_idx = col_idx
            if cell.value == "Adj Close":
                adj_close_col_idx = col_idx

        # 如果列名行没有 Close/Adj Close 列，则追加
        if close_col_idx is None:
            # 找到最后一列
            last_col = ws.max_column
            ws.cell(row=target_row_idx, column=last_col + 1, value=close_val)
            print(f"  {ticker}: 写入 Close={close_val}（新列）")
        else:
            ws.cell(row=target_row_idx, column=close_col_idx, value=close_val)
            print(f"  {ticker}: 写入 Close={close_val}")

        if adj_close_col_idx is not None:
            ws.cell(row=target_row_idx, column=adj_close_col_idx, value=close_val)

    wb.save(price_file)
    print(f"\n✅ 收盘价补全完成，已保存至: {price_file}")

    # 打印摘要
    print(f"\n补全摘要：")
    for ticker, price in close_prices.items():
        print(f"  {ticker}: {price}")


if __name__ == "__main__":
    main()
